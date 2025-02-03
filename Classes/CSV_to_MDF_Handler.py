import logging
from openpyxl import load_workbook

from asammdf import MDF, Signal
import numpy as np
import pandas as pd

class CSV_to_MDF_Handler:
    def __init__(self):
        pass

    @staticmethod
    def replace_cell(stringList, find_array, replace_array):
        outCell = []
        workCell = []
        for elem in stringList:
            workCell.append(elem)

        row_num = 0
        while row_num < len(workCell):
            if row_num <= (len(workCell) - len(find_array)):
                match_counter = 0
                for find_array_index in range(0, len(find_array)):
                    if workCell[row_num + find_array_index] == find_array[find_array_index]:
                        match_counter += 1
                if match_counter == len(find_array):
                    # print("====")
                    # print(workCell[row_num-1])
                    # print(workCell[row_num])
                    # print(workCell[row_num+1])
                    # print("====")
                    # print(len(find_array))
                    for replace_array_index in range(0, len(replace_array)):
                        outCell.append(replace_array[replace_array_index])
                    row_num += len(find_array)
                else:
                    outCell.append(workCell[row_num])
                    row_num += 1
            else:
                outCell.append(workCell[row_num])
                row_num += 1

        workCell = []
        for elem in outCell:
            workCell.append(elem)

        return outCell

    def exec_substitution3(self, input_file_path, output_file_path):
        """ This is a quick summary line used as a description of the object.
        quick summary line used as a description of the object
        """
        print("hello there")
        return
        wb_in = load_workbook(input_file_path)
        ws_in = wb_in["input_file_sheet"]

        print("import input file : DONE")

        wb_find_replace = load_workbook("find_replace_file_path")
        ws_find_replace = wb_find_replace["find_replace_file_sheet"]

        print("open find replace file : DONE")

        return
        substitution_list_from_excel = []
        for row in ws_find_replace.iter_rows(min_row=2, min_col=1, max_col=2):
            single_substitution = {"find": "string", "replace": "string"}
            for colNum, cell in enumerate(row):
                if isinstance(cell.value, str):
                    if colNum == 0:
                        single_substitution["find"] = cell.value.split("\n")
                    elif colNum == 1:
                        single_substitution["replace"] = cell.value.split("\n")
            substitution_list_from_excel.append(single_substitution)

        print(substitution_list_from_excel)
        print("import find replace file : DONE")

        wb_in.save(filename=output_file_path)
        wbOut = load_workbook(output_file_path)
        wsOut = wbOut["input_file_sheet"]

        print("saved Copy of Input file: DONE")

        columnPreconditionIndex = 0
        columnActionIndex = 1
        columnExpectedResIndex = 2

        for substitution in substitution_list_from_excel:
            find_array = substitution['find']
            replace_array = substitution['replace']

            for col in wsOut.iter_cols(min_row=1, min_col=columnPreconditionIndex,
                                       max_col=columnExpectedResIndex):
                for rowNum, cell in enumerate(col):
                    # tmp_array = []
                    # print(type(cellString))
                    if isinstance(cell.value, str):
                        cellString = cell.value
                        # print(len(cellString))
                        splitvalue = cellString.split('\n')
                        newCellValue = self.replace_cell(splitvalue, find_array, replace_array)
                        cellNewValueString = '\n'.join(newCellValue)
                        # print(len(cellNewValueString))
                        # print(cellNewValueString)
                        cell.value = str(cellNewValueString)
                        # print(cellNewValueString)
                        # tmp_array.append(row + "\n")
                        # print(tmp_array)
                        # Write in rich text
                        # optsheet.write_rich_string('A1', red, splitvalue[0], splitvalue[1])
                        # for elem in newCellValue:
                        # print("==================")
                        # optsheet.write_rich_string(rowNum + 1, 2, *tmp_array)
                    # Split characters
                    # print(cell.value)
            # optbook.close()

        print("substitution : DONE")
        wbOut.save(filename=output_file_path)

        print("file output saving : DONE")

    def exec_cleanup(self):

        wb_in = load_workbook(self.cfg_data.input_file_path)
        ws_in = wb_in[self.cfg_data.input_file_sheet]

        print("import input file : DONE")
        wb_in.save(filename=self.cfg_data.output_file_path)

        wbOut = load_workbook(self.cfg_data.output_file_path)
        wsOut = wbOut[self.cfg_data.input_file_sheet]

        print("saved Copy of Input file: DONE")

        columnPreconditionIndex = 0
        columnExpectedResIndex = 1

        for col in wsOut.iter_cols(min_row=1, max_row=1640,
                                   min_col=columnPreconditionIndex,
                                   max_col=columnExpectedResIndex):
            for cell in col:
                if isinstance(cell.value, str):
                    cellString = cell.value
                    cell.value = str(cellString)

        print("cancellazioni righe vuote : DONE")

        wbOut.save(filename=self.cfg_data.output_file_path)

        print("file output saving : DONE")

    def exec_bullet_lists_fix(self):

        wb_in = load_workbook(self.cfg_data.input_file_path)
        ws_in = wb_in[self.cfg_data.input_file_sheet]

        print("import input file : DONE")
        wb_in.save(filename=self.cfg_data.output_file_path)

        wbOut = load_workbook(self.cfg_data.output_file_path)
        wsOut = wbOut[self.cfg_data.input_file_sheet]

        print("saved Copy of Input file: DONE")

        columnPreconditionIndex = 0
        columnExpectedResIndex = 1

        for col in wsOut.iter_cols(min_row=1, max_row=1640,
                                   min_col=columnPreconditionIndex,
                                   max_col=columnExpectedResIndex):
            for cell in col:
                if isinstance(cell.value, str):
                    cellString = cell.value
                    cell.value = str(cellString)

        print("sistemazione elenchi puntati : DONE")

        wbOut.save(filename=self.cfg_data.output_file_path)

        print("file output saving : DONE")

    def exec_polarion_to_excel_converter(self):

        wb_in = load_workbook(self.cfg_data.input_file_path)
        ws_in = wb_in[self.cfg_data.input_file_sheet]

        print("import input file : DONE")
        wb_in.save(filename=self.cfg_data.output_file_path)

        wbOut = load_workbook(self.cfg_data.output_file_path)
        wsOut = wbOut[self.cfg_data.input_file_sheet]

        print("saved Copy of Input file: DONE")

        columnPreconditionIndex = 0
        columnExpectedResIndex = 1

        for col in wsOut.iter_cols(min_row=1, max_row=1640,
                                   min_col=columnPreconditionIndex,
                                   max_col=columnExpectedResIndex):
            for cell in col:
                if isinstance(cell.value, str):
                    cellString = cell.value
                    #cellNewValueString = 0
                    cell.value = str(cellString)

        print("polarion to excel conversion: DONE")

        wbOut.save(filename=self.cfg_data.output_file_path)

        print("file output saving : DONE")

    def exec_excel_to_polarion_converter(self):

        wb_in = load_workbook(self.cfg_data.input_file_path)
        ws_in = wb_in[self.cfg_data.input_file_sheet]

        print("import input file : DONE")
        wb_in.save(filename=self.cfg_data.output_file_path)

        wbOut = load_workbook(self.cfg_data.output_file_path)
        wsOut = wbOut[self.cfg_data.input_file_sheet]

        print("saved Copy of Input file: DONE")

        columnPreconditionIndex = 0
        columnExpectedResIndex = 1
        logging.debug("max row 1640.... why?")

        for col in wsOut.iter_cols(min_row=1, max_row=1640,
                                   min_col=columnPreconditionIndex,
                                   max_col=columnExpectedResIndex):
            for cell in col:
                if isinstance(cell.value, str):
                    cellString = cell.value
                    cell.value = str(cellString)

        print("excel to polarion conversion: DONE")

        wbOut.save(filename=self.cfg_data.output_file_path)

        print("file output saving : DONE")

    # -------------------
    # START CONFIGURATION
    # -------------------
    @staticmethod
    def exec_substitution(input_file_path, use_same_input_file_name, output_file_name):
        path = "C:\\Users\\stefano.fortunati\\Documents\\_LAVORO\\Schaeffler\\PXL_EDAG\\"

        # FILENAME SENZA ESTENSIONE !!!
        filename = "19.08.2024_Schaeffler1_L012"
        #df = pd.read_csv(path + filename + ".txt", sep="\t", encoding="latin1")
        print(input_file_path)
        df = pd.read_csv(input_file_path, sep="\t", encoding="latin1")

        # -------------------
        # END CONFIGURATION
        # -------------------

        # Inserimento di una colonna "time" con valori in 'datetime'
        df.insert(1, 'time', pd.to_datetime(df[df.columns[0]],format="%d.%m.%Y %H:%M:%S,%f"))
        start_time = df['time'].iloc[0]

        # Inserimento di una colonna "relative_time" con valori in 'datetime'
        df.insert(2, 'relative_time', df['time'] - start_time)

        # removed first column DateTime
        # Cambia i formati dei dati in ingresso, escludendo la colonna DateTime
        columns_to_process = df.columns[3:]

        for col_name in columns_to_process:
            if df[col_name].dtypes == 'object':
                try:
                    df[col_name] = df[col_name].str.replace(',', '.').astype(float)
                except:
                    pass
            elif df[col_name].dtypes == 'int64':
                pass
                # print("int ", col_name)
            elif df[col_name].dtypes == 'float64':
                # print("float64 ", col_name)
                df[col_name] = df[col_name].replace(',', '.').astype(float)
            else:
                print(col_name)
                print(df[col_name].dtypes)


        # ############
        # STEP1 : modificare il data set, per trasformare da str a decimali i numeri
        # ############
        # original_data_frame = ""

        column_to_delete = []
        for col_name in columns_to_process:
            number_of_nan_values = df[col_name].isnull().sum()
            column_size = df[col_name].size
            if number_of_nan_values == column_size:
                # print(col_name , ": all Nan")
                column_to_delete.append(col_name)

        columns_to_process = [col for col in columns_to_process if col not in column_to_delete]
        #############
        ## STEP2 : creare asse dei tempi
        #############
        timestamps = np.array(df["relative_time"].apply(lambda x: x.seconds))

        #############
        ## STEP2 : creare i segnali effettivi dal dataframe
        #############

        signals_list = []

        for col_name in columns_to_process:
            signal = Signal(samples=np.array(df[col_name], dtype=df[col_name].dtypes),
                            timestamps=timestamps, name=col_name, unit='')

            signals_list.append(signal)

        # create empty MDf version 4.00 file
        with (MDF(version='4.10') as mdf4):
            # append the signals to the new file
            mdf4.append(signals_list, comment='imported')
            mdf4.start_time = start_time.to_pydatetime()  # datetime.fromisoformat("2024-08-06 17:00:00")
            # save new file
            if use_same_input_file_name == True:
                out_filename = input_file_path[:-4] + ".mf4"
            else:
                out_filename = output_file_name
                mdf4.save(path + filename + ".mf4", overwrite=True)
            #print(out_filename)
            mdf4.save(out_filename, overwrite=True)
