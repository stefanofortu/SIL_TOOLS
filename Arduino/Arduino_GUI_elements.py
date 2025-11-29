import os
import pandas
import pandas as pd
from PySide6.QtCore import Qt
from PySide6.QtGui import QPainter, QColor
from PySide6.QtWidgets import QMessageBox, QGroupBox, QHBoxLayout, QLabel, QSpinBox, QPushButton, QSpacerItem, \
    QSizePolicy, QWidget
from openpyxl.reader.excel import load_workbook


#from Arduino.PWM_Reader_Widget import Pwm_Reader_Widget

class TestSequence(pd.DataFrame):
    def __init__(self):
        super().__init__()

    def load_from_file(self, file_path, verbose=False):
        if not os.path.exists(file_path):
            raise FileNotFoundError

        try:
            # Carica il file Excel
            wb = load_workbook(file_path)  # percorso del file
            sheet = wb.active  # usa il primo foglio attivo
        except Exception as e:
            raise Exception(e)

        # Leggi il valore della cella A1
        first_row_parameter = sheet["A1"].value
        first_row_value = sheet["B1"].value
        first_row_delay = sheet["C1"].value

        # Controlla prima cella
        if (first_row_parameter != "Parameter" or
                first_row_value != "Value" or
                first_row_delay != "Delay[s]"):
            raise IndentationError("Il file Excel non è formattato correttamente")
        else:
            super().__init__(columns=["Parameter", "Value", "Delay[s]"])
        # Aggiorna label con il nome del file
        data = []  # lista per contenere tutte le righe non vuote

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Controlla se la riga contiene almeno un valore
            if any(cell is not None for cell in row):
                parameter = row[0] if len(row) > 1 else ""
                value = row[1] if len(row) > 2 else ""
                delay_s = row[2]
                self.loc[len(self)] = [parameter, value, delay_s]

        if verbose:
            # Stampa tutte le righe lette (include l'intestazione)
            print(self)


class PumpRowGroupBox(QGroupBox):
    def __init__(self, name, pump_id, serial_communication):
        super(PumpRowGroupBox, self).__init__()
        self.pump_id = pump_id
        self.setTitle(name)
        self.group_box_serial_communication = serial_communication

        self.setStyleSheet("""
        QGroupBox::title {
           subcontrol-origin: margin; padding: 0 5px; color: darkblue;
            }
        QGroupBox {
            font-size: 14px;
            font-weight: bold;
            border: 1px solid blue;
            border-radius: 5px;
            margin-top: 16px;
            }
        """)
        group_layout = QHBoxLayout()

        group_layout.addWidget(QLabel("Speed"))

        self.spin_box_pump_speed = QSpinBox(self)
        self.spin_box_pump_speed.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        self.spin_box_pump_speed.setMinimumWidth(75)
        self.spin_box_pump_speed.setMaximum(100)
        self.spin_box_pump_speed.setSuffix(" %")
        group_layout.addWidget(self.spin_box_pump_speed)

        self.set_speed_button = QPushButton(self)
        self.set_speed_button.setText("Set ➡")
        self.set_speed_button.setStyleSheet("font-weight: bold;")
        self.set_speed_button.setMaximumWidth(75)
        self.set_speed_button.pressed.connect(
            lambda: self.set_pwm_speed(pump_id=self.pump_id,
                                       speed_value=self.spin_box_pump_speed.value())
        )

        group_layout.addWidget(self.set_speed_button)

        # Add explandable spacer
        group_layout.addStretch()

        # Add fixed-width spacer
        group_layout.addItem(QSpacerItem(20, 20, QSizePolicy.Fixed, QSizePolicy.Minimum))

        group_layout.addWidget(QLabel("Status :"))

        self.label_pump_status = QLabel("xxxx ms ()")
        self.label_pump_status.setStyleSheet("background-color: white; padding: 5px;")
        group_layout.addWidget(self.label_pump_status)

        # Add small circle widget
        self.circle_widget_pump = CircleWidget(diameter=18, color="red")
        group_layout.addWidget(self.circle_widget_pump)
        self.circle_widget_pump.setColor("orange")

        # Add another fixed-width spacer
        group_layout.addItem(QSpacerItem(20, 20, QSizePolicy.Fixed, QSizePolicy.Minimum))

        self.setLayout(group_layout)

    def set_pwm_speed(self, pump_id, speed_value):
        print("==================================")
        print("========SET DUTY===================")
        duty_cycle_parameter_offset = 16
        parameter_id = duty_cycle_parameter_offset + pump_id
        message = f"4:{parameter_id:02X}:{speed_value:02X}"
        self.group_box_serial_communication.write_message(message=message, verbose=True)

    def disable_manual_input(self, arg: bool):
        self.spin_box_pump_speed.setDisabled(arg)
        self.set_speed_button.setDisabled(arg)

    def enable_manual_input(self, arg: bool):
        self.spin_box_pump_speed.setEnabled(arg)
        self.set_speed_button.setEnabled(arg)

    def update_pwm_speed(self, value_hex):
        value_dec = int(value_hex, 16)
        self.spin_box_pump_speed.setValue(value_dec)

    def update_pump_status(self, value_hex):
        value_dec = int(value_hex, 16)
        value_dec = value_dec % 2500
        pump_status = ""
        if 450 <= value_dec <= 550:
            pump_status = "FEEDBACK OK"
            self.circle_widget_pump.setColor("green")
        elif 950 <= value_dec <= 1050:
            pump_status = "DRY RUN"
            self.circle_widget_pump.setColor("red")
        elif 1450 <= value_dec <= 1550:
            pump_status = "OVERTEMPERATURE"
            self.circle_widget_pump.setColor("red")
        elif 1950 <= value_dec <= 2050:
            pump_status = "UNDER-OVER-VOLTAGE"
            self.circle_widget_pump.setColor("red")
        else:
            pump_status = "UNKNOWN"
            self.circle_widget_pump.setColor("orange")

        self.label_pump_status.setText(f"{value_dec} ms ({pump_status})")


class CircleWidget(QWidget):
    def __init__(self, diameter=20, color="red"):
        super().__init__()
        self.diameter = diameter
        self.color = color
        self.setFixedSize(diameter, diameter)  # ensure the widget stays square

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setBrush(QColor(self.color))
        painter.setPen(Qt.NoPen)
        painter.drawEllipse(0, 0, self.diameter, self.diameter)

    def setColor(self, new_color):
        self.color = QColor(new_color)
        self.update()  # Trigger repaint


class SwitchButton(QPushButton):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCheckable(True)
        self.setChecked(True)
        self.update_style()
        self.toggled.connect(self.update_style)

    def update_style(self):
        print("here")
        if self.isChecked():
            self.setText("MANUAL MODE")
            self.setStyleSheet("""
                QPushButton {
                    background-color: #ff3131;
                    color: white;
                    border-radius: 14px;
                    padding: 6px 12px;
                }
            """)
        else:
            self.setText("AUTOMATIC MODE")
            self.setStyleSheet("""
                QPushButton {
                    background-color: #4cd964;
                    color: white;
                    border-radius: 14px;
                    padding: 6px 12px;
                }
            """)
        #self.automatic_mode_is_active()

    def automatic_mode_is_active(self):
        if self.isChecked():
            print("Automatic mode active")
            return True
        else:
            print("Automatic mode not active")
            return False


class Automatic_Mode_Button(QPushButton):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCheckable(True)
        self.setChecked(False)
        self.update_style()
        self.toggled.connect(self.update_style)

    def update_style(self):
        if self.isChecked():
            self.setText("AUTOMATIC MODE")
            self.setStyleSheet("""
                QPushButton {
                    background-color: #4cd964;
                    color: white;
                    border-radius: 14px;
                    padding: 6px 12px;
                }
            """)
        else:
            self.setText("MANUAL MODE")
            self.setStyleSheet("""
                QPushButton {
                    background-color: #ff3131;
                    color: white;
                    border-radius: 14px;
                    padding: 6px 12px;
                }
            """)

    def automatic_mode_active(self):
        return True if self.isChecked() else False
