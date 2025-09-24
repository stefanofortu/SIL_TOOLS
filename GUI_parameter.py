# =================================================
# *                    HEADER
# =================================================
# Industrie Saleri Italo spa
# Via Ruca 406
# 25065 Lumezzane (Bs), Italia
# E-mail: mirko.gozio@saleri.it
# =================================================
# *                    HISTORY
# =================================================
# Date: 07/03/2023     Developer: Mirko Gozio
# v01: first release
#
# PYTHON script
#
import os
import sys
import time
import shutil
import imageio
import datetime
import string
import math
import scipy
from scipy import stats
import matplotlib
import openpyxl
import random as random
import matplotlib.pyplot as plt
from matplotlib import cm
from matplotlib.ticker import LinearLocator, FormatStrFormatter
from matplotlib.pyplot import *
from matplotlib import colors
from matplotlib.font_manager import FontProperties
import numpy as np
import pandas as pd
import csv
from mpl_toolkits.mplot3d import Axes3D
from scipy.interpolate import griddata
import seaborn as sns
from datetime import timedelta
import tkinter as tk
from tkinter import filedialog


import platform



def post_parameter():
    pump_name = entry_pump_name.get()
    working_dir = var_wd.get()
    voltage_string = entry_voltage.get()
    speed_string = entry_speed.get()
    speed_udm_string = speed_udm
    deltatime_speed = float(entry_deltaTime_speed.get())
    avg_time = float(entry_avgTime.get())
    tol_dVdT = float(entry_tolerance.get())
    pumpOff_time = float(entry_pumpOff.get())
    tolWarning = float(entry_tolWarning.get())
    if ";" in voltage_string:
        sep_voltage = ";"
    elif "," in voltage_string:
        sep_voltage = ","
    elif ":" in voltage_string:
        sep_voltage = ":"
    V2test = []
    for item in voltage_string.split(sep_voltage): V2test.append(float(item))
    if ";" in speed_string:
        sep_speed = ";"
    elif "," in speed_string:
        sep_speed = ","
    elif ":" in speed_string:
        sep_speed = ":"
    speed2test = []
    for item in speed_string.split(sep_speed): speed2test.append(float(item))
    os.system("cls")
    print("\n\n*******************")
    print("Pump: " + pump_name)
    print("Working directory: " + working_dir)
    print("Tested voltages: " + str(V2test))
    print("Tested speeds: " + str(speed2test) + "," + speed_udm_string)
    print("Stationary time for each speed: " + str(deltatime_speed) + " s")
    print("Averaging time: " + str(avg_time) + " s")
    print("Tolerance dV/dt: " + str(tol_dVdT))
    print("Pump off time: " + str(pumpOff_time) + " s")
    print("Tolerance warning: " + str(tolWarning) + "%")
    # vecchio script
    pd.set_option('display.max_rows', None)  # Nessun limite sul numero di righe
    pd.set_option('display.max_columns', None)  # Nessun limite sul numero di colonne
    pd.set_option('display.width', None)  # Larghezza illimitata per evitare l'andata a capo
    pd.set_option('display.max_colwidth', None)  # Mostra il contenuto completo delle celle
    df_output = pd.DataFrame()
    list_files = []
    if ".csv" in working_dir:
        list_files.append(working_dir)
    else:
        for fil in os.listdir(working_dir):
            list_files.append(working_dir + "/" + fil)
    print(list_files)
    for fil in list_files:
        if ".csv" in fil and "output" not in fil:
            name_pump = fil.split("_")[-1].split(".")[0]  # numero della pompa
            df_tmp = pd.read_csv(fil, decimal=".", sep=",", encoding="latin1", parse_dates=["Time"],
                                 date_parser=custom_date_parser_logger)
            time_delta = timedelta(seconds=1)
            if "Delta P - P2" in df_tmp.columns or "Delta P - P1" in df_tmp.columns:
                if df_tmp["Delta P - P2"].max() - df_tmp["Delta P - P2"].min() > df_tmp["Delta P - P1"].max() - df_tmp[
                    "Delta P - P1"].min():
                    df_tmp["DeltaP"] = df_tmp["Delta P - P2"]
                elif df_tmp["Delta P - P1"].max() - df_tmp["Delta P - P1"].min() > df_tmp["Delta P - P2"].max() - \
                        df_tmp["Delta P - P2"].min():
                    df_tmp["DeltaP"] = df_tmp["Delta P - P1"]
            df_tmp['DVpump_Dt'] = df_tmp['Vpump'].diff() / df_tmp['Time'].diff().dt.total_seconds()
            df_tmp['DIpump_Dt'] = df_tmp['Ipump'].diff() / df_tmp['Time'].diff().dt.total_seconds()
            df_tmp['DQ_Dt'] = df_tmp['Q'].diff() / df_tmp['Time'].diff().dt.total_seconds()
            df_tmp['DdP_Dt'] = df_tmp['DeltaP'].diff() / df_tmp['Time'].diff().dt.total_seconds()
            list_time = search_change_voltage(df_tmp, tol_dVdT, time_delta)
            list_time_filtered = []
            for time in list_time:
                Vpump_i = df_tmp[df_tmp["Time"] == time - timedelta(seconds=0.2)]["Vpump"].unique()[0]
                DVpump_Dt_i = df_tmp[df_tmp["Time"] == time - timedelta(seconds=0.2)]["DVpump_Dt"].unique()[0]
                # print(time, Vpump_i, DVpump_Dt_i)
                if Vpump_i <= df_tmp[df_tmp["Time"] == time + timedelta(seconds=0.4)]["Vpump"].unique()[
                    0] + 0.1 and Vpump_i >= df_tmp[df_tmp["Time"] == time + timedelta(seconds=0.4)]["Vpump"].unique()[
                    0] - 0.1: continue
                min_val = 10000
                for el in V2test:
                    if abs(el - Vpump_i) <= min_val:
                        min_val = abs(el - Vpump_i)
                    else:
                        min_val = min_val
                if min_val < 0.1:
                    list_time_filtered.append(time)
                else:
                    continue
            fig1, axs1 = plt.subplots(2, 1, figsize=(15, 10))
            axs1[0].plot(df_tmp["Time"], df_tmp["DVpump_Dt"], label="dVpump/dt")
            axs1[0].plot(df_tmp["Time"], df_tmp["DIpump_Dt"], label="dIpump/dt")
            axs1[0].plot(df_tmp["Time"], df_tmp["Vpump"], label="Vpump")
            axs1[0].plot(df_tmp["Time"], df_tmp["Ipump"], label="Ipump")
            axs1[1].plot(df_tmp["Time"], df_tmp["Q"], label="Q")
            axs1[1].plot(df_tmp["Time"], df_tmp["DeltaP"], label="DeltaP")
            axs1[1].plot(df_tmp["Time"], df_tmp["Vpump"], label="Vpump")
            axs1[1].plot(df_tmp["Time"], df_tmp["Ipump"], label="Ipump")
            fig1.suptitle(pump_name + " " + name_pump)
            axs1[0].legend()
            axs1[1].legend()
            for el in list_time_filtered:
                axs1[0].axvline(x=el, color='black', linestyle='--')
                axs1[1].axvline(x=el, color='black', linestyle='--')
                for ii, speed in enumerate(speed2test[::-1]):
                    start = el - timedelta(seconds=pumpOff_time + avg_time + ii * deltatime_speed)
                    end = el - timedelta(seconds=pumpOff_time + ii * deltatime_speed)
                    axs1[1].axvline(x=start, color='gray', linestyle='--')
                    axs1[1].axvline(x=end, color="gray", linestyle=":")
            plt.show()
            if len(list_time_filtered) != len(V2test):
                print("ERRORE FILTRAGGIO TIME")
            else:
                for time in list_time_filtered:
                    for ii, speed_i in enumerate(speed2test[::-1]):
                        plot_chart = False
                        df_filt = df_tmp[(df_tmp["Time"] >= time - timedelta(
                            seconds=pumpOff_time + avg_time + ii * deltatime_speed)) & (
                                                     df_tmp["Time"] <= time - timedelta(
                                                 seconds=pumpOff_time + ii * deltatime_speed))].copy()
                        df_filt["Vpump"] = df_filt["Vpump"].round(2)
                        df_filt["Ipump"] = df_filt["Ipump"].round(3)
                        df_filt["Q"] = df_filt["Q"].round(3)
                        df_filt["DeltaP"] = df_filt["DeltaP"].round(3)
                        df_filt["Tair"] = df_filt["Tair"].round(2)
                        df_filt["Tcoolant"] = df_filt["Tcoolant"].round(2)
                        df_2mean = df_filt[["Vpump", "Ipump", "Q", "DeltaP", "Tair", "Tcoolant"]]
                        df_output_tmp = df_2mean.mean().to_frame().T
                        df_output_tmp["Pump"] = name_pump
                        df_output_tmp["speed"] = speed_i
                        df_output_tmp["Vpump"] = df_output_tmp["Vpump"].round(2)
                        df_output_tmp["Ipump"] = df_output_tmp["Ipump"].round(3)
                        df_output_tmp["Q"] = df_output_tmp["Q"].round(3)
                        df_output_tmp["DeltaP"] = df_output_tmp["DeltaP"].round(3)
                        df_output_tmp["Tair"] = df_output_tmp["Tair"].round(2)
                        df_output_tmp["Tcoolant"] = df_output_tmp["Tcoolant"].round(2)
                        print("## Pump: " + name_pump + ", " + str(
                            round(df_output_tmp["Vpump"].unique()[0], 0)) + "V, " + str(speed_i) + "rpm ##")
                        plot_chart, V_warning = check_variazione(df_filt["Vpump"], df_output_tmp["Vpump"].unique()[0],
                                                                 "Vpump", plot_chart, tolWarning)
                        plot_chart, I_warning = check_variazione(df_filt["Ipump"], df_output_tmp["Ipump"].unique()[0],
                                                                 "Ipump", plot_chart, tolWarning)
                        plot_chart, DeltaP_warning = check_variazione(df_filt["DeltaP"],
                                                                      df_output_tmp["DeltaP"].unique()[0], "DeltaP",
                                                                      plot_chart, tolWarning)
                        plot_chart, Q_warning = check_variazione(df_filt["Q"], df_output_tmp["Q"].unique()[0], "Q",
                                                                 plot_chart, tolWarning)
                        df_output_tmp["V warning"] = V_warning
                        df_output_tmp["I warning"] = I_warning
                        df_output_tmp["DeltaP warning"] = DeltaP_warning
                        df_output_tmp["Q warning"] = Q_warning
                        if plot_chart == True:
                            fig, axs = plt.subplots(2, 2, figsize=(10, 10))
                            axs[0, 0].plot(df_filt["Time"], df_filt["Vpump"])
                            axs[0, 0].axhline(y=df_output_tmp["Vpump"].unique()[0], color="green", ls="--")
                            axs[0, 1].plot(df_filt["Time"], df_filt["Ipump"])
                            axs[0, 1].axhline(y=df_output_tmp["Ipump"].unique()[0], color="green", ls="--")
                            axs[1, 0].plot(df_filt["Time"], df_filt["DeltaP"])
                            axs[1, 0].axhline(y=df_output_tmp["DeltaP"].unique()[0], color="green", ls="--")
                            axs[1, 1].plot(df_filt["Time"], df_filt["Q"])
                            axs[1, 1].axhline(y=df_output_tmp["Q"].unique()[0], color="green", ls="--")
                            fig.suptitle(pump_name + " " + name_pump + ", " + str(
                                round(df_output_tmp["Vpump"].unique()[0], 0)) + "V, " + str(speed_i) + " " + str(
                                speed_udm_string))
                            axs[0, 0].set_xlabel("Time")
                            axs[0, 1].set_xlabel("Time")
                            axs[1, 0].set_xlabel("Time")
                            axs[1, 1].set_xlabel("Time")
                            axs[0, 0].set_ylabel("Vpump [V]")
                            axs[0, 1].set_ylabel("Ipump [A]")
                            axs[1, 0].set_ylabel("DeltaP [bar]")
                            axs[1, 1].set_ylabel("Q [lpm]")
                            axs[0, 0].set_ylim(
                                df_output_tmp["Vpump"].unique()[0] - 0.1 * df_output_tmp["Vpump"].unique()[0],
                                df_output_tmp["Vpump"].unique()[0] + 0.1 * df_output_tmp["Vpump"].unique()[0])
                            axs[0, 1].set_ylim(
                                df_output_tmp["Ipump"].unique()[0] - 0.1 * df_output_tmp["Ipump"].unique()[0],
                                df_output_tmp["Ipump"].unique()[0] + 0.1 * df_output_tmp["Ipump"].unique()[0])
                            axs[1, 0].set_ylim(
                                df_output_tmp["DeltaP"].unique()[0] - 0.1 * df_output_tmp["DeltaP"].unique()[0],
                                df_output_tmp["DeltaP"].unique()[0] + 0.1 * df_output_tmp["DeltaP"].unique()[0])
                            axs[1, 1].set_ylim(df_output_tmp["Q"].unique()[0] - 0.1 * df_output_tmp["Q"].unique()[0],
                                               df_output_tmp["Q"].unique()[0] + 0.1 * df_output_tmp["Q"].unique()[0])
                            plt.tight_layout()
                            plt.show()
                        print("")
                        df_output = pd.concat([df_output, df_output_tmp], ignore_index=True)
    df_output["Efficienza"] = 100 * (5 / 3 * df_output["DeltaP"] * df_output["Q"]) / (
                df_output["Vpump"] * df_output["Ipump"])
    df_output["Efficienza"] = df_output["Efficienza"].round(3)
    df_output["Vtarget"] = df_output["Vpump"].round().astype(int)
    print(df_output)
    if ".csv" not in working_dir:
        working_dir = working_dir
    else:
        working_dir = os.path.dirname(working_dir)
    os.makedirs(working_dir + "/output_csv", exist_ok=True)
    with pd.ExcelWriter(working_dir + "/" + pump_name + ".xlsx", engine='openpyxl') as writer:
        for pump in df_output["Pump"].unique():
            df_output_filt = df_output[df_output["Pump"] == pump]

            priority_map = {12: 1, 9: 2, 16: 3}

            df_output_filt = df_output_filt.sort_values(
                by=["Vtarget", "speed"],
                key=lambda col: col.map(priority_map) if col.name == "Vtarget" else col,
                ascending=[True, True]
            )


            #df_output_filt = df_output_filt.sort_values(by=['Vtarget', 'speed'], ascending=[True, True])
            df_output_filt = df_output_filt[
                ["Pump", "Vtarget", "speed", "Vpump", "Ipump", "Q", "DeltaP", "Efficienza", "Tcoolant", "Tair",
                 "V warning", "I warning", "DeltaP warning", "Q warning"]]
            df_output_filt.to_csv(working_dir + "/output_csv/" + pump_name + "_" + pump + "_output.csv", index=False,
                             decimal=",", sep=";")
            sheet_name = pump
            df_output_filt.to_excel(writer, sheet_name=sheet_name, index=False)

            from openpyxl import load_workbook
            import platform
            if platform.system() == "Windows":
                path_separator = "\\"
            elif platform.system() == "Linux":
                path_separator = "/"
            else:
                path_separator = "\\"

            pump_name = df_output_filt["Pump"].iloc[0]
            filename_in = "C:\\Users\\stefano.fortunati\\Desktop\\Tamb_xxx.xlsx"
            filename_out = filename_in.replace("_xxx.xlsx", "_" + pump_name + ".xlsx")

            #filename_out = "C:\\Users\\stefano.fortunati\\Desktop\\Tamb_P01.xlsx"
            try:
                os.makedirs(os.path.dirname(filename_out), exist_ok=True)
                shutil.copy(filename_in, filename_out)
            except shutil.SameFileError:
                print("Source and destination represents the same file.")
            except PermissionError:
                print("Permission denied.")
                # except:
                # logger.error("Error occurred while copying file.")

            wb = load_workbook("C:\\Users\\stefano.fortunati\\Desktop\\Tamb_P01.xlsx")

            # OPTION 2 to modify an existing wb
            ws = wb["Data_Entry"]
            ws.cell(row=2, column=1).value = pump_name # df_output_filt["Pump"].iloc[0]

            v_target_list = df_output_filt["Vtarget"].tolist()
            for i, val in enumerate(v_target_list, start=2):
                ws.cell(row=i, column=22, value=val)

            v_pump_list = df_output_filt["Vpump"].tolist()
            for i, val in enumerate(v_pump_list, start=2):
                ws.cell(row=i, column=5, value=val)

            i_pump_list = df_output_filt["Ipump"].tolist()
            for i, val in enumerate(i_pump_list, start=2):
                ws.cell(row=i, column=6, value=val)

            i_pump_list = df_output_filt["Q"].tolist()
            for i, val in enumerate(i_pump_list, start=2):
                ws.cell(row=i, column=7, value=val)

            i_pump_list = df_output_filt["DeltaP"].tolist()
            for i, val in enumerate(i_pump_list, start=2):
                ws.cell(row=i, column=8, value=val)

            i_pump_list = df_output_filt["Tcoolant"].tolist()
            for i, val in enumerate(i_pump_list, start=2):
                ws.cell(row=i, column=10, value=val)

            i_pump_list = df_output_filt["Tair"].tolist()
            for i, val in enumerate(i_pump_list, start=2):
                ws.cell(row=i, column=11, value=val)
            #df_output_filt["Vtarget"]
            #"speed", "Vpump", "Ipump", "Q", "DeltaP", "Efficienza", "Tcoolant", "Tair",
            # "V warning", "I warning", "DeltaP warning", "Q warning"]
            #self.move_sheet_tab_to_end()
            #self.delete_all_other_sheets()

            #self.ws.sheet_properties.outlinePr.summaryBelow = False

            #ws.cell(row_num, category_column).value = "Categories"
            #ws.cell(row_num, in_column, "in")
            #ws.cell(row_num, savings_in_column, "savings in")
            #ws.cell(row_num, out_column, "out")
            #ws.cell(row_num, savings_out_column, "savings out")
            #ws.cell(row_num, no_tags_column, "no tags")

            # Insert data starting at row 10, col 2 (B10)
            #ws.cell(row=10, column=7).value= "Inserted Value"

            wb.save(filename_out)
            wb.close()


def custom_date_parser_logger(date_string):
    return pd.to_datetime(date_string, format="%d/%m/%Y %H:%M:%S.%f")


def search_change_voltage(df, val, time_delta):
    time_DVpump_switch = df[(df["DVpump_Dt"] > val) | (df["DVpump_Dt"] < -val)]
    time_DVpump_switch["diff_seconds"] = time_DVpump_switch["Time"].diff()
    time_DVpump_switch = time_DVpump_switch.dropna(subset=["diff_seconds"])
    time_DVpump_switch = time_DVpump_switch[time_DVpump_switch["diff_seconds"] >= pd.Timedelta(time_delta)]
    time_list = time_DVpump_switch["Time"].tolist()
    return (time_list)


def check_variazione(df, mean_value, stringa, plot_chart, tol):
    plot_chart_ = plot_chart
    range_col = df.max() - df.min()
    if df.min == 0:
        print(stringa + " minima = 0, test fallito")
        plot_chart_ = True
        warning = "Critical"
    else:
        if range_col > 0.01 * tol * mean_value:
            print(f"range_col {range_col} tol {tol} mean_value {mean_value}")
            print(f"{stringa} range (max:{df.max()}, min:{df.min()}) maggiore del {tol} % della media ({mean_value}), valutare i risultati con attenzione")
            plot_chart_ = True
            warning = "Warning"
        else:
            # print(stringa + " range minore del " + str(tol) + "% della media")
            warning = "Ok"
    return plot_chart_, warning


def wd_selection():
    root = tk.Tk()
    if var_filefolder.get() == "FOLDER":
        wrk_path = filedialog.askdirectory(title="Working directory")
        if wrk_path:
            var_wd.set(wrk_path)
            root.destroy()
    elif var_filefolder.get() == "FILE":
        wrk_path = filedialog.askopenfilename(title="Select file")
        if wrk_path:
            var_wd.set(wrk_path)
            root.destroy()


def update_speed_udm(selected_option):
    # Assegna il valore alla variabile speed_udm in base alla selezione
    global speed_udm
    speed_udm = selected_option
    # print(f"Selected speed unit: {speed_udm}")


def update_filefolder(selected_option):
    # Assegna il valore alla variabile speed_udm in base alla selezione
    global filefolder_
    filefolder_ = selected_option
    button_wd.config(text=filefolder_)
    print(f"Selected speed unit: {filefolder_}")


def toggle_labels():
    # Controlla se il flag è selezionato
    if flag_var.get():
        # Aggiungi i nuovi label se il flag è attivo
        label_tolerance_value.grid(row=8, column=0, padx=10, pady=10, sticky='w')
        label_pumpOff_time.grid(row=9, column=0, padx=10, pady=10, sticky='w')
        label_tolWarning.grid(row=10, column=0, padx=10, pady=10, sticky='w')
        entry_tolerance.grid(row=8, column=1, padx=10, pady=10)
        entry_pumpOff.grid(row=9, column=1, padx=10, pady=10)
        entry_tolWarning.grid(row=10, column=1, padx=10, pady=10)
    else:
        # Rimuovi i nuovi label se il flag è disattivato
        label_tolerance_value.grid_forget()
        label_pumpOff_time.grid_forget()
        label_tolWarning.grid_forget()
        entry_tolerance.grid_forget()
        entry_pumpOff.grid_forget()
        entry_tolWarning.grid_forget()


def import_data():
    root = tk.Tk()
    import_data_filepath = filedialog.askopenfilename(title="Import data")
    f = open(import_data_filepath, "r")
    testo = f.readlines()
    f.close()
    root.destroy()
    for line in testo:
        if "pump name" in line.lower():
            var_pump_name.set(line.split(":")[1].replace("\n", ""))
        elif "tested voltages" in line.lower():
            var_voltage.set(line.split(":")[1].replace("\n", ""))
        elif "tested speeds" in line.lower():
            var_speed.set(line.split(":")[1].replace("\n", ""))
        elif "dt speeds" in line.lower():
            var_deltaTime_speed.set(line.split(":")[1].replace("\n", ""))
        elif "avg time" in line.lower():
            var_avgTime.set(line.split(":")[1].replace("\n", ""))
        elif "tolerance value" in line.lower():
            var_tolerance.set(line.split(":")[1].replace("\n", ""))
        elif "pump-off time" in line.lower():
            var_pumpOff.set(line.split(":")[1].replace("\n", ""))
        elif "tolerance warning" in line.lower():
            var_tolWarning.set(line.split(":")[1].replace("\n", ""))


def export_importdatafile():
    testo = "Pump Name: " + str(var_pump_name.get()) + "\nTested Voltages: " + str(
        var_voltage.get()) + "\nTested Speeds: " + str(var_speed.get()) + "\ndt speeds: " + str(
        var_deltaTime_speed.get()) + "\nAvg time: " + str(var_avgTime.get()) + "\nTolerance value: " + str(
        var_tolerance.get()) + "\nPump-off time: " + str(var_pumpOff.get()) + "\nTolerance warning: " + str(
        var_tolWarning.get())
    root = tk.Tk()
    wrk_path = filedialog.askdirectory(title="Export import data file")
    root.destroy()
    if wrk_path:
        # Ensure the file path is correct, use os.path.join for cross-platform compatibility
        file_path = os.path.join(wrk_path, "importdataexample.dat")
        # Writing to the file using a context manager (with statement)
        with open(file_path, "w") as f:
            f.writelines(testo)


def Help_():
    root_help = tk.Tk()
    root_help.title("Help")
    label = tk.Label(root_help, text="Nobody can help you", font=("Arial", 24))
    label.grid(row=1, column=1, padx=20, pady=20)
    label2 = tk.Label(root_help,
                      text="But if you really need you can write to mirko.gozio@saleri.com, hourly rate 50€/h",
                      font=("Arial", 14))
    label2.grid(row=2, column=1, padx=20, pady=20)


root = tk.Tk()
root.title("POST-PARAMETER SALERI")
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_width = screen_width // 2
window_height = screen_height // 1
root.geometry(f"{window_width}x{window_height}")
root.configure(bg='lightblue')
label_filefolder = tk.Label(root, text="Work on", bg="lightblue")
label_filefolder.grid(row=0, column=0, padx=10, pady=10, sticky="e")
options_filefolder = ["FILE", "FOLDER"]
var_filefolder = tk.StringVar()
var_filefolder.set(options_filefolder[1])
dropdown_filefolder = tk.OptionMenu(root, var_filefolder, *options_filefolder, command=update_filefolder)
dropdown_filefolder.grid(row=0, column=1, padx=10, pady=10)
label_pump_name = tk.Label(root, text="Insert pump name", bg="lightblue")
label_pump_name.grid(row=1, column=0, padx=10, pady=10, sticky='e')  # Posizioniamo la label in (0, 0)
# Creiamo un'entry (campo di input) in cui l'utente può scrivere una stringa
var_pump_name = tk.StringVar()
entry_pump_name = tk.Entry(root, textvariable=var_pump_name)
entry_pump_name.grid(row=1, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
# Creiamo label per selezionare la working directory
label_wd = tk.Label(root, text="Working directory/File", bg="lightblue")
label_wd.grid(row=2, column=0, padx=10, pady=10, sticky='e')
# Creiamo variabile dinamica per la stringa della working directory
var_wd = tk.StringVar()
# Creiamo label per visualizzare working directory
label_var_wd = tk.Label(root, textvariable=var_wd, wraplength=400, bg="lightblue")
label_var_wd.grid(row=2, column=1, columnspan=2, pady=10)
# Creiamo button per un pulsante per selezionare la working directory
button_wd = tk.Button(root, text=var_filefolder.get(), command=lambda: [wd_selection()])
button_wd.grid(row=2, column=3, columnspan=2, pady=10)
# Creiamo label per definire i voltaggi testati
label_Voltage = tk.Label(root, text="Tested voltages", bg="lightblue")
label_Voltage.grid(row=3, column=0, padx=10, pady=10, sticky='e')
# Creiamo variabile dinamica per la stringa della working directory
var_voltage = tk.StringVar()
# Creiamo un'entry (campo di input) in cui l'utente può scrivere una stringa
entry_voltage = tk.Entry(root, textvariable=var_voltage)
entry_voltage.grid(row=3, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
# Creiamo label per definire i voltaggi testati
label_Voltage_info = tk.Label(root, text="*Insert ; between each value", bg="lightblue")
label_Voltage_info.grid(row=3, column=3, padx=10, pady=10, sticky='e')
# Creiamo label per definire i voltaggi testati
label_Speed = tk.Label(root, text="Tested speeds", bg="lightblue")
label_Speed.grid(row=4, column=0, padx=10, pady=10, sticky='e')
# Creiamo variabile dinamica per la stringa della working directory
var_speed = tk.StringVar()
# Creiamo un'entry (campo di input) in cui l'utente può scrivere una stringa
entry_speed = tk.Entry(root, textvariable=var_speed)
entry_speed.grid(row=4, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
# Creiamo un option menù per le unità di misura della velocità
speed_udm = ""
options = ["rpm", "pwm", "lin"]
var_speed_udm = tk.StringVar()
var_speed_udm.set("")
dropdown_speed = tk.OptionMenu(root, var_speed_udm, *options, command=update_speed_udm)
dropdown_speed.grid(row=4, column=2, padx=10, pady=10)
# Creiamo label per definire i voltaggi testati
label_Speed_info = tk.Label(root, text="*Insert ; between each value. Insert the speeds in the order they were tested",
                            bg="lightblue", wraplength=200)
label_Speed_info.grid(row=4, column=3, padx=10, pady=10, sticky='e')
# Creiamo label per definire i voltaggi testati
label_deltaTime_speed = tk.Label(root, text="dt speed [s]", bg="lightblue")
label_deltaTime_speed.grid(row=5, column=0, padx=10, pady=10, sticky='e')
# Creiamo variabile dinamica per la stringa della working directory
var_deltaTime_speed = tk.StringVar()
# Creiamo un'entry (campo di input) in cui l'utente può scrivere una stringa
entry_deltaTime_speed = tk.Entry(root, textvariable=var_deltaTime_speed)
entry_deltaTime_speed.grid(row=5, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
label_deltaTime_speed_info = tk.Label(root, text="*Stationary time for each speed", bg="lightblue")
label_deltaTime_speed_info.grid(row=5, column=3, padx=10, pady=10, sticky='e')
# Creiamo label per definire i voltaggi testati
label_avgTime = tk.Label(root, text="Avg time [s]", bg="lightblue")
label_avgTime.grid(row=6, column=0, padx=10, pady=10, sticky='e')
# Creiamo variabile dinamica per la stringa della working directory
var_avgTime = tk.StringVar()
# Creiamo un'entry (campo di input) in cui l'utente può scrivere una stringa
entry_avgTime = tk.Entry(root, textvariable=var_avgTime)
entry_avgTime.grid(row=6, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
label_avgTime_info = tk.Label(root, text="*Average time for each speed", bg="lightblue")
label_avgTime_info.grid(row=6, column=3, padx=10, pady=10, sticky='e')
# Creiamo un flag variable
flag_var = tk.BooleanVar()
# Creiamo un checkbox (Checkbutton) per attivare/disattivare i label
check_button = tk.Checkbutton(root, text="More Options", variable=flag_var, command=toggle_labels, bg="lightblue")
check_button.grid(row=7, column=0, padx=10, pady=10)
# Nuovi label da mostrare quando il checkbox è selezionato
label_tolerance_value = tk.Label(root, text="Tolerance value dV/dt", bg="lightblue")
label_pumpOff_time = tk.Label(root, text="Pump off time [s]", bg="lightblue")
label_tolWarning = tk.Label(root, text="Tolerance warning %", bg="lightblue")
var_tolerance = tk.StringVar()
var_tolerance.set("5")
var_pumpOff = tk.StringVar()
var_pumpOff.set("7")
var_tolWarning = tk.StringVar()
var_tolWarning.set(2.5)
entry_tolerance = tk.Entry(root, textvariable=var_tolerance)
entry_pumpOff = tk.Entry(root, textvariable=var_pumpOff)
entry_tolWarning = tk.Entry(root, textvariable=var_tolWarning)
# Creiamo un comando di import data file
button1 = tk.Button(root, text="Import data", relief="ridge", command=lambda: [import_data()])
button1.grid(row=0, column=5, columnspan=2, pady=10)
# Creiamo un comando di export import data
button2 = tk.Button(root, text="Export Import data file", relief="ridge", command=lambda: [export_importdatafile()])
button2.grid(row=1, column=5, columnspan=2, pady=10)
# Creiamo un pulsante per lanciare il post-processing
button = tk.Button(root, text="EXECUTE", command=lambda: [post_parameter()], width=20, height=3, font=("Arial", 14))
button.grid(row=7, column=3, columnspan=2, pady=10)
# Creiamo un pulsante help
button = tk.Button(root, text="Help", command=lambda: [Help_()])
button.grid(row=11, column=0, columnspan=2, pady=10)

# Avviamo il loop principale della finestra
root.mainloop()


