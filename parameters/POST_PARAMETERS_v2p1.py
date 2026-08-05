# =================================================
# *                    HEADER
# =================================================
# Industrie Saleri Italo spa
# Via Ruca 406
# 25065 Lumezzane (Bs), Italia
# E-mail: mirko.gozio@saleri.it
# =================================================
# *                    HISTORY
# =================================================
# Date: 07/03/2023     Developer: Mirko Gozio
# v01: first release
#
# PYTHON script
#
import os
import sys
import time
import shutil
import datetime
import math
import scipy
from scipy import stats
import matplotlib
import openpyxl
import random as random
import matplotlib.pyplot as plt
from matplotlib import cm
from matplotlib.ticker import LinearLocator, FormatStrFormatter
from matplotlib.pyplot import *
from matplotlib import colors
from matplotlib.font_manager import FontProperties
import numpy as np
import pandas as pd
import csv
from mpl_toolkits.mplot3d import Axes3D
from scipy.interpolate import griddata
from datetime import timedelta
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk,messagebox
import platform
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from libraries.DW_to_Py_Converter import Dewesoft_Converter

def post_parameter():

# lettura variabili inserite e verifica che nessuna sia mancante

    pump_name=entry_pump_name.get()
    files2read = [f.strip() for f in filesread.get().split(";") if f.strip()]
    folder_path=selected_folder.get() #cartella in cui ci sono i file
    xlsx_folder=find_xlsx_folder(folder_path) #cartella in cui ci sono gli excel di output
    prj_folder=find_project_folder(folder_path)
    output_folder = os.path.join(prj_folder, "elaborazione")
# Crea la cartella solo se non esiste
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    if not filesread.get().strip():
        messagebox.showerror("Error", "Select at least one file")
        return
    voltages2plot=var_voltage.get().split(";")
    if not var_voltage.get().strip():
        messagebox.showerror("Error", "Insert at least one voltage")
        return
    speed_udm=var_speed_udm.get()
    if speed_udm=="":
        messagebox.showerror("Error", "Define speed unit")
        return
    speed2plot=var_speed.get().split(";")
    if not var_speed.get().strip():
        messagebox.showerror("Error", "Insert at least one speed")
        return
    deltatime_speed=var_deltaTime_speed.get()
    if deltatime_speed=="":
        messagebox.showerror("Error", "Define delta time speed")
        return
    avgTime=var_avgTime.get()
    if avgTime=="":
        messagebox.showerror("Error", "Define average time")
        return
    tol_dVdT=var_tolerance.get()
    if tol_dVdT=="":
        messagebox.showerror("Error", "Define tolerance")
        return
    pumpOff_time=var_pumpOff.get()
    if pumpOff_time=="":
        messagebox.showerror("Error", "Define pump-off time")
        return
    tolWarning=var_tolWarning.get()
    if tolWarning=="":
        messagebox.showerror("Error", "Define tolerance warning threashold")
        return  
    #print("\n\n*******************")
    #print("Pump: " + pump_name)
    #print("Files to read: " + files2read)
    #print("Tested voltages: " + voltages2plot)
    #print("Tested speeds: " + str(speed2plot) + "," + speed_udm)
    #print("Stationary time for each speed: " + str(deltatime_speed) + " s")
    #print("Averaging time: " + str(avgTime) + " s")
    #print("Tolerance dV/dt: " + str(tol_dVdT))
    #print("Pump off time: " + str(pumpOff_time) + " s")
    #print("Tolerance warning: " + str(tolWarning) + "%")

# copio i file da leggere in locale al fine di ridurre i tempi I/O

    sistema=platform.system()

    if sistema == "Windows":
        base_tmp = os.getenv('TEMP')
        tmp_dir = os.path.join(base_tmp, "tmp_postparameter__")
    else:
        tmp_dir = os.path.join(os.getcwd(), ".tmp_postparameter__")

    # Se la cartella esiste, eliminata prima di crearla
    if os.path.exists(tmp_dir):
        try:
            shutil.rmtree(tmp_dir)
        except PermissionError:
            messagebox.showerror("Error", f"Cannot remove existing temporary folder: {tmp_dir}")
            return None

    # Crea la cartella
    os.makedirs(tmp_dir)
    
    # Imposta hidden su Windows
    if sistema == "Windows":
        try:
            import ctypes
            FILE_ATTRIBUTE_HIDDEN = 0x02
            ctypes.windll.kernel32.SetFileAttributesW(tmp_dir, FILE_ATTRIBUTE_HIDDEN)
        except PermissionError:
            print("Non è stato possibile rendere la cartella nascosta (permessi insufficienti).")

    files_copied = [shutil.copy(fil, tmp_dir) for fil in files2read]
    filesread_ = tuple(files_copied)



    #pd.set_option('display.max_rows', None)  # Nessun limite sul numero di righe
    pd.set_option('display.max_columns', None)  # Nessun limite sul numero di colonne
    pd.set_option('display.width', None)  # Larghezza illimitata per evitare l'andata a capo
    pd.set_option('display.max_colwidth', None)  # Mostra il contenuto completo delle celle
    df_output = pd.DataFrame()
    list_files = []

# inizio lettura file
    plot_tolerance_warning=False
    for fil in filesread_:
       print("\n\n \t *****")
       id_pump=fil.split("_")[-1].split(".")[0]
       if ".csv" in fil:
          df_tmp = safe_read_csv(fil,decimal=".",sep=",",encoding="latin1",parse_dates=["Time"],date_parser=custom_date_parser_logger)
       if ".dxd" in fil:
          df_tmp = read_dxd(fil)
       time_delta = timedelta(seconds=1)
       if "Delta P - P2" in df_tmp.columns or "Delta P - P1" in df_tmp.columns:
          if df_tmp["Delta P - P2"].max() - df_tmp["Delta P - P2"].min() > df_tmp["Delta P - P1"].max() - df_tmp["Delta P - P1"].min():
             df_tmp["DeltaP"] = df_tmp["Delta P - P2"] #2
             print("selected P2")
          elif df_tmp["Delta P - P1"].max() - df_tmp["Delta P - P1"].min() > df_tmp["Delta P - P2"].max() - df_tmp["Delta P - P2"].min():
             df_tmp["DeltaP"] = df_tmp["Delta P - P1"] #1
             print("selected P1")

       df_tmp['DVpump_Dt'] = df_tmp['Vpump'].diff() / df_tmp['Time'].diff().dt.total_seconds()
       df_tmp['DIpump_Dt'] = df_tmp['Ipump'].diff() / df_tmp['Time'].diff().dt.total_seconds()
       df_tmp['DQ_Dt'] = df_tmp['Q'].diff() / df_tmp['Time'].diff().dt.total_seconds()
       df_tmp['DdP_Dt'] = df_tmp['DeltaP'].diff() / df_tmp['Time'].diff().dt.total_seconds()
       list_time = search_change_voltage(df_tmp, float(tol_dVdT), time_delta)
       
       list_time_filtered = []
       for time in list_time:
           Vpump_i = df_tmp[df_tmp["Time"] == time - timedelta(seconds=0.2)]["Vpump"].unique()[0]
           DVpump_Dt_i = df_tmp[df_tmp["Time"] == time - timedelta(seconds=0.2)]["DVpump_Dt"].unique()[0]
             # print(time, Vpump_i, DVpump_Dt_i)
           if Vpump_i <= df_tmp[df_tmp["Time"] == time + timedelta(seconds=0.4)]["Vpump"].unique()[0] + 0.1 and Vpump_i >= df_tmp[df_tmp["Time"] == time + timedelta(seconds=0.4)]["Vpump"].unique()[0] - 0.1: continue
           min_val = 10000
           for el in voltages2plot:
               el=float(el)
               if abs(el - Vpump_i) <= min_val:
                   min_val = abs(el - Vpump_i)
               else:
                   min_val = min_val
           if min_val < 0.1:
               list_time_filtered.append(time)
           else:
                 continue
       
       fig, ax = plt.subplots(figsize=(15, 6))  # una sola figura e un solo asse

# --- Plotta solo le curve principali sul secondo asse ---
       ax.plot(df_tmp["Time"], df_tmp["Q"], label="Q")
       ax.plot(df_tmp["Time"], df_tmp["DeltaP"], label="DeltaP")
       ax.plot(df_tmp["Time"], df_tmp["Vpump"], label="Vpump")
       ax.plot(df_tmp["Time"], df_tmp["Ipump"], label="Ipump")
       ax.set_ylim(0,)
       ax.grid(ls="--")

# --- Titolo e legenda ---
       ax.set_title(pump_name + " " + id_pump)
       ax.legend()

# --- Linee verticali ---
       for el in list_time_filtered:
          ax.axvline(x=el, color='black', linestyle='--')
          for ii, speed in enumerate(speed2plot[::-1]):
             speed = float(speed)
             start = el - timedelta(seconds=float(pumpOff_time) + float(avgTime) + ii * float(deltatime_speed))
             end = el - timedelta(seconds=float(pumpOff_time) + ii * float(deltatime_speed))
             ax.axvline(x=start, color='gray', linestyle='--')
             ax.axvline(x=end, color='gray', linestyle=':')

# --- Mostra il grafico ---
       pumpOff_time, avgTime, deltatime_speed, stop_flag,failed_note,plot_tolerance_warning=show_dynamic_plot(df_tmp, list_time_filtered, speed2plot, float(pumpOff_time), float(avgTime), float(deltatime_speed),plot_tolerance_warning,os.path.basename(fil))

       if stop_flag:break
# --- Completato check grafico

       if len(list_time_filtered) != len(voltages2plot):
           messagebox.showerror("Error", "Error on filtering time")
           break
       else:
           for time in list_time_filtered:
              for ii, speed_i in enumerate(speed2plot[::-1]):
                  plot_chart=False
                  df_filt = df_tmp[(df_tmp["Time"] >= time - timedelta(seconds=pumpOff_time + avgTime + ii * deltatime_speed)) & (df_tmp["Time"] <= time - timedelta(seconds=pumpOff_time + ii * deltatime_speed))].copy()
                  df_filt["Vpump"] = df_filt["Vpump"].round(2)
                  df_filt["Ipump"] = df_filt["Ipump"].round(3)
                  df_filt["Q"] = df_filt["Q"].round(3)
                  df_filt["DeltaP"] = df_filt["DeltaP"].round(3)
                  df_filt["Tair"] = df_filt["Tair"].round(2)
                  df_filt["Tcoolant"] = df_filt["Tcoolant"].round(2)
                  df_2mean = df_filt[["Vpump", "Ipump", "Q", "DeltaP", "Tair", "Tcoolant"]]
                  df_output_tmp = df_2mean.mean().to_frame().T
                  df_output_tmp["Pump"] = pump_name+"_"+id_pump
                  df_output_tmp["speed"] = speed_i
                  df_output_tmp["Vpump"] = df_output_tmp["Vpump"].round(2)
                  df_output_tmp["Ipump"] = df_output_tmp["Ipump"].round(3)
                  df_output_tmp["Q"] = df_output_tmp["Q"].round(3)
                  df_output_tmp["DeltaP"] = df_output_tmp["DeltaP"].round(3)
                  df_output_tmp["Tair"] = df_output_tmp["Tair"].round(2)
                  df_output_tmp["Tcoolant"] = df_output_tmp["Tcoolant"].round(2)
                  df_output_tmp["Note"] = failed_note

                  print("\n## Pump: " + id_pump + ", " + str(round(df_output_tmp["Vpump"].unique()[0], 0)) + "V, " + str(speed_i) + speed_udm)
                  plot_chart, V_warning = check_variazione(df_filt["Vpump"], df_output_tmp["Vpump"].unique()[0],"Vpump", plot_chart, float(tolWarning))
                  plot_chart, I_warning = check_variazione(df_filt["Ipump"], df_output_tmp["Ipump"].unique()[0],"Ipump", plot_chart, float(tolWarning))
                  plot_chart, DeltaP_warning = check_variazione(df_filt["DeltaP"],df_output_tmp["DeltaP"].unique()[0], "DeltaP",plot_chart, float(tolWarning))
                  plot_chart, Q_warning = check_variazione(df_filt["Q"], df_output_tmp["Q"].unique()[0], "Q",plot_chart, float(tolWarning))
                  df_output_tmp["V warning"] = V_warning
                  df_output_tmp["I warning"] = I_warning
                  df_output_tmp["DeltaP warning"] = DeltaP_warning
                  df_output_tmp["Q warning"] = Q_warning
                 
                  if plot_chart == True and plot_tolerance_warning==True:
                      plt.close("all")
                      fig_, axs_ = plt.subplots(2, 2, figsize=(10, 10))
                      axs_[0, 0].plot(df_filt["Time"], df_filt["Vpump"])
                      axs_[0, 0].axhline(y=df_output_tmp["Vpump"].unique()[0], color="green", ls="--")

                      axs_[0, 1].plot(df_filt["Time"], df_filt["Ipump"])
                      axs_[0, 1].axhline(y=df_output_tmp["Ipump"].unique()[0], color="green", ls="--")
                      axs_[1, 0].plot(df_filt["Time"], df_filt["DeltaP"])
                      axs_[1, 0].axhline(y=df_output_tmp["DeltaP"].unique()[0], color="green", ls="--")
                      axs_[1, 1].plot(df_filt["Time"], df_filt["Q"])
                      axs_[1, 1].axhline(y=df_output_tmp["Q"].unique()[0], color="green", ls="--")
                      fig_.suptitle(pump_name + " " + id_pump + ", " + str(round(df_output_tmp["Vpump"].unique()[0], 0)) + "V, " + str(speed_i) + " " +speed_udm)
                      axs_[0, 0].set_xlabel("Time")
                      axs_[0, 1].set_xlabel("Time")
                      axs_[1, 0].set_xlabel("Time")
                      axs_[1, 1].set_xlabel("Time")
                      axs_[0, 0].set_ylabel("Vpump [V]")
                      axs_[0, 1].set_ylabel("Ipump [A]")
                      axs_[1, 0].set_ylabel("DeltaP [bar]")
                      axs_[1, 1].set_ylabel("Q [lpm]")
                      axs_[0, 0].set_ylim(
                      df_output_tmp["Vpump"].unique()[0] - 0.1 * df_output_tmp["Vpump"].unique()[0],
                      df_output_tmp["Vpump"].unique()[0] + 0.1 * df_output_tmp["Vpump"].unique()[0])
                      axs_[0, 1].set_ylim(df_output_tmp["Ipump"].unique()[0] - 0.1 * df_output_tmp["Ipump"].unique()[0],df_output_tmp["Ipump"].unique()[0] + 0.1 * df_output_tmp["Ipump"].unique()[0])
                      axs_[1, 0].set_ylim(df_output_tmp["DeltaP"].unique()[0] - 0.1 * df_output_tmp["DeltaP"].unique()[0],df_output_tmp["DeltaP"].unique()[0] + 0.1 * df_output_tmp["DeltaP"].unique()[0])
                      axs_[1, 1].set_ylim(df_output_tmp["Q"].unique()[0] - 0.1 * df_output_tmp["Q"].unique()[0],df_output_tmp["Q"].unique()[0] + 0.1 * df_output_tmp["Q"].unique()[0])
                      fig_.tight_layout()
                      plt.show()
                  df_output = pd.concat([df_output, df_output_tmp], ignore_index=True)
    df_output["Efficienza"] = 100 * (5 / 3 * df_output["DeltaP"] * df_output["Q"]) / (df_output["Vpump"] * df_output["Ipump"])
    df_output["Efficienza"] = df_output["Efficienza"].round(3)
    df_output["Vtarget"] = df_output["Vpump"].round().astype(int)
    #print(df_output)

# elimina la cartella temporanea creata
    if os.path.exists(tmp_dir):
        try:
            shutil.rmtree(tmp_dir)
        except PermissionError:
            messagebox.showerror("Error", f"Cannot remove existing temporary folder: {tmp_dir}")
            return None


# Definisco le categorie
    categories = {"<0": lambda x: x < 0,"0-50": lambda x: (x >= 0) & (x <= 50),">50": lambda x: x > 50}

    for pump in df_output["Pump"].unique():
        print("df_output[Pump] ", pump)
        df_filt=df_output[df_output["Pump"]==pump]
        #print(df_filt)

 # Dizionario per salvare i dati filtrati solo se esistono
        filtered_data = {}
        for cat_name, condition in categories.items():
            df_cat = df_filt[condition(df_filt["Tcoolant"])]
            voltages_order = [float(v) for v in voltages2plot]
            # Applica l'ordinamento personalizzato per Vtarget e crescente per speed
            df2export = df_cat.copy()
            df2export["Vtarget"] = pd.Categorical(df2export["Vtarget"], categories=voltages_order, ordered=True)
            df2export = df2export.sort_values(by=["Vtarget", "speed"], ascending=[True, True]).reset_index(drop=True)
            print(df2export)
            if not df2export.empty:
                source_file = excel_file_name.get()# os.path.join(xlsx_folder, "Tamb_xxx.xlsx")
                #new_name = "Tamb_" + pump.strip() + ".xlsx"
                destination_file = source_file.replace("xxx",pump.strip())
                #    = os.path.join(output_folder, new_name)
                print(f"source_file", source_file)
                print(f"destination_file", destination_file)
                shutil.copy2(source_file, destination_file)
                compile_excel(df2export, destination_file, speed_udm)

                destination_folder = os.path.dirname(destination_file)
                #print(destination_folder)
                destination_folder = destination_folder.replace("/", "\\")
                #destination_folder = destination_folder + "\\"
                print(f"destination_folder", destination_folder)
                os.startfile(destination_folder)
                '''
                if cat_name=="0-50":
                    source_file=os.path.join(xlsx_folder, "Tamb_xxx.xlsx")
                    new_name="Tamb_"+pump.strip()+".xlsx"
                    destination_file=os.path.join(output_folder, new_name)
                    print(f"source_file",source_file)
                    print(f"destination_file",destination_file)
                    shutil.copy2(source_file, destination_file)
                    compile_excel(df2export,destination_file,speed_udm)
                    #import os
                    os.startfile(output_folder)
                elif cat_name==">50":
                    source_file=os.path.join(xlsx_folder, "Tmax_xxx.xlsx")
                    new_name="Tmax_"+pump.strip()+".xlsx"
                    destination_file=os.path.join(output_folder, new_name)
                    shutil.copy2(source_file, destination_file)
                    print(f"source_file",source_file)
                    print(f"destination_file",destination_file)
                    compile_excel(df2export,destination_file,speed_udm)
                elif cat_name=="<0":
                    source_file=os.path.join(xlsx_folder, "Tmin_xxx.xlsx")
                    new_name="Tmin_"+pump.strip()+".xlsx"
                    destination_file=os.path.join(output_folder, new_name)
                    shutil.copy2(source_file, destination_file)
                    print(f"source_file",source_file)
                    print(f"destination_file",destination_file)
                    compile_excel(df2export,destination_file,speed_udm)
               '''
def read_dxd(filename):
    dewesoft_converter = Dewesoft_Converter()
    dewesoft_converter.open(filename=filename)
    df_tmp = dewesoft_converter.to_pandas(verbose=False)
    dewesoft_converter.close()
    df_tmp["Time"]=df_tmp["timestamp"]
    df_tmp["Time"] = df_tmp["Time"].dt.tz_localize(None).dt.floor("ms")
    df_tmp=df_tmp.drop(["timestamp","Time[s]"],axis=1)
    for col in df_tmp.columns:
        if "elet" in col.lower() and "elta" not in col.lower():
            df_tmp=df_tmp.drop(col,axis=1)
        elif "idr" in col.lower():
            df_tmp=df_tmp.drop(col,axis=1)
        elif "eff" in col.lower():
            df_tmp=df_tmp.drop(col,axis=1)
    return df_tmp


def compile_excel(df,fil,speed_udm):
    wb=load_workbook(fil)
    ws=wb["Data_Entry"]
    offset=1
    ws.cell(row=2+offset,column=1+offset).value=df["Pump"].unique()[0]
    #ws.cell(row=1,column=4).value="Target Speed "+speed_udm
    
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # verde vivo
    fill_red   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # rosso vivo
    fill_orange = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")  # arancione vivo

    #v_target_list=df["Vtarget"].tolist()
    #for i,val in enumerate(v_target_list,start=2):
    #    ws.cell(row=i+offset,column=3,value=val)

    v_pump_list=df["Vpump"].tolist()
    for i,val in enumerate(v_pump_list,start=2):
        ws.cell(row=i+offset,column=5+offset,value=val)


    i_pump_list=df["Ipump"].tolist()
    for i,val in enumerate(i_pump_list,start=2):
        ws.cell(row=i+offset,column=6+offset,value=val)

    q_pump_list=df["Q"].tolist()
    for i,val in enumerate(q_pump_list,start=2):
        ws.cell(row=i+offset,column=7+offset,value=val)

    dp_pump_list=df["DeltaP"].tolist()
    for i,val in enumerate(dp_pump_list,start=2):
        ws.cell(row=i+offset,column=8+offset,value=val)

    #speed_pump_list=df["speed"].tolist()
    #for i,val in enumerate(speed_pump_list,start=2):
    #    ws.cell(row=i+offset,column=4,value=val)

    Tcoolant_pump_list=df["Tcoolant"].tolist()
    for i,val in enumerate(Tcoolant_pump_list,start=2):
        ws.cell(row=i+offset,column=10+offset,value=val)

    Tair_pump_list=df["Tair"].tolist()
    for i,val in enumerate(Tair_pump_list,start=2):
        ws.cell(row=i+offset,column=11+offset,value=val)

    Vwarning_pump_list=df["V warning"].tolist()
    for i,val in enumerate(Vwarning_pump_list,start=2):
        ws.cell(row=i+offset,column=14+offset,value=val)
        cell=ws.cell(row=i+offset,column=14+offset)
        if val=="Ok":
           cell.fill=fill_green
        else:
           cell.fill=fill_orange

    Iwarning_pump_list=df["I warning"].tolist()
    for i,val in enumerate(Iwarning_pump_list,start=2):
        ws.cell(row=i+offset,column=15+offset,value=val)
        cell=ws.cell(row=i+offset,column=15+offset)
        if val=="Ok":
           cell.fill=fill_green
        else:
           cell.fill=fill_orange

    Qwarning_pump_list=df["Q warning"].tolist()
    for i,val in enumerate(Qwarning_pump_list,start=2):
        ws.cell(row=i+offset,column=16+offset,value=val)
        cell=ws.cell(row=i+offset,column=16+offset)
        if val=="Ok":
           cell.fill=fill_green
        else:
           cell.fill=fill_orange

    dPwarning_pump_list=df["DeltaP warning"].tolist()
    for i,val in enumerate(dPwarning_pump_list,start=2):
        ws.cell(row=i+offset,column=17+offset,value=val)
        cell=ws.cell(row=i+offset,column=17+offset)
        if val=="Ok":
           cell.fill=fill_green
        else:
           cell.fill=fill_orange

    FailedNote_pump_list=df["Note"].tolist()
    for i,val in enumerate(FailedNote_pump_list,start=2):
        ws.cell(row=i+offset,column=18+offset,value=val)
        cell=ws.cell(row=i+offset,column=18+offset)
        if val!="":
           cell.fill=fill_red
        else:
           cell.fill=fill_green

    wb.save(fil)
    wb.close()                   


    
def find_project_folder(start_path):
    """
    Risale di un livello alla volta da start_path finché non trova
    una cartella con nome strutturato come AAAA_MM_GG_NNNN (es. 2025_01_01_0040).
    Se non la trova entro 5 livelli, cerca una cartella con 10 cifre consecutive nel nome.

    Restituisce:
        path della cartella trovata oppure start_path se non trovata.
    """
    current_dir = os.path.abspath(start_path)
    pattern_date = re.compile(r"^\d{4}_\d{2}_\d{2}_\d{4}$")
    pattern_digits = re.compile(r"^\d{10}$")
    max_levels = 6

    # Primo tentativo: cerca formato AAAA_MM_GG_NNNN
    for _ in range(max_levels):
        folder_name = os.path.basename(current_dir)
        if pattern_date.match(folder_name):
            return current_dir
        parent_dir = os.path.dirname(current_dir)
        if parent_dir == current_dir:  # raggiunta la root
            break
        current_dir = parent_dir

    # Secondo tentativo: cerca nome con 10 cifre consecutive
    current_dir = os.path.abspath(start_path)
    for _ in range(max_levels):
        folder_name = os.path.basename(current_dir)
        if pattern_digits.match(folder_name):
            return current_dir
        parent_dir = os.path.dirname(current_dir)
        if parent_dir == current_dir:
            break
        current_dir = parent_dir

    # Nessuna cartella trovata
    return start_path
      
def find_xlsx_folder(start_path):
    """
    Risale di un livello alla volta a partire da start_path
    e cerca la prima cartella che contenga almeno un file
    'Tamb_*.xlsx', 'Tmax_*.xlsx' o 'Tmin_*.xlsx'.

    Restituisce:
        folder_path oppure None
    """
    current_dir = os.path.abspath(start_path)
    pattern = re.compile(r"^(Tamb_|Tmax_|Tmin_)xxx\.xlsx$", re.IGNORECASE)

    while True:
        # Cerca ricorsivamente in questa directory
        for root, _, files in os.walk(current_dir):
            if any(pattern.match(f) for f in files):
                return root  # Restituisce solo la cartella che contiene i file trovati

        # Risale di un livello
        parent_dir = os.path.dirname(current_dir)
        if parent_dir == current_dir:  # raggiunta la root → stop
            break
        current_dir = parent_dir

    return None



def safe_read_csv(path, **kwargs):
    """Legge un CSV saltando eventuali righe iniziali vuote."""
    with open(path, "r", encoding=kwargs.get("encoding", "latin1")) as f:
        lines = f.readlines()

    # Rimuove eventuali righe iniziali vuote o solo con separatori
    clean_lines = []
    for line in lines:
        if line.strip():  # tiene solo righe non vuote
            clean_lines = lines[lines.index(line):]
            break

    if not clean_lines:
        raise ValueError(f"File {path} is empty or invalid")

    # Legge il CSV dai dati "puliti"
    from io import StringIO
    return pd.read_csv(StringIO("".join(clean_lines)), **kwargs)



def show_dynamic_plot(df_tmp, list_time_filtered, speed2plot, pumpOff_time, avgTime, deltatime_speed,plot_tolerance_warning,fil):
    # Crea nuova finestra Tk
    win = tk.Toplevel()
    win.title(fil)
    win.geometry("1300x700")

    # --- Figure matplotlib ---
    fig, ax = plt.subplots(figsize=(10,5))
    canvas = FigureCanvasTkAgg(fig, master=win)
    canvas.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # --- Disegno base ---
    ax.plot(df_tmp["Time"], df_tmp["Q"], label="Q")
    ax.plot(df_tmp["Time"], df_tmp["DeltaP"], label="DeltaP")
    ax.plot(df_tmp["Time"], df_tmp["Vpump"], label="Vpump")
    ax.plot(df_tmp["Time"], df_tmp["Ipump"], label="Ipump")
    ax.legend()
    ax.grid(ls="--")

    stop_flag = tk.BooleanVar(value=False)
    test_failed_note = tk.StringVar(value="")

    def update_lines(*args):
    # Se una Entry è temporaneamente vuota → non aggiornare
        if not var_pumpOff.get() or not var_avgTime.get() or not var_deltatime_speed.get():
            return

        try:
            pOff = float(var_pumpOff.get())
            aTime = float(var_avgTime.get())
            dtSpeed = float(var_deltatime_speed.get())
        except ValueError:
            return  # ignora input non numerici momentanei

    # Rimuovi solo le linee verticali precedenti (quelle con label='axv')
        for l in ax.get_lines():
            if l.get_label() == 'axv':
                l.remove()

    # Ridisegna le linee aggiornate
        for el in list_time_filtered:
            ax.axvline(x=el, color='black', linestyle='--', label='axv')
            for ii, speed in enumerate(speed2plot[::-1]):
                start = el - timedelta(seconds=pOff + aTime + ii * dtSpeed)
                end = el - timedelta(seconds=pOff + ii * dtSpeed)
                ax.axvline(x=start, color='gray', linestyle='--', label='axv')
                ax.axvline(x=end, color='gray', linestyle=':', label='axv')

        canvas.draw_idle()

    # --- Funzione di chiusura ---
    def close_window():
        win.destroy()  # Chiude la finestra e sblocca win.wait_window()

    def block_process():
        stop_flag.set(True)  # segnala stop
        win.destroy()         # chiude e interrompe ciclo

    def test_failed():
        # Crea una finestra modale per la nota
        note_win = tk.Toplevel(win)
        note_win.title("Test Failed - Inserisci nota")
        note_win.geometry("400x200")
        note_win.geometry("400x200")
        note_win.configure(bg="lightblue")

        tk.Label(note_win, text="Write a note about the failed test:",bg="lightblue").pack(pady=10)
        text_box = tk.Text(note_win, height=5, width=40)
        text_box.pack(padx=10, pady=5)

        def confirm_note():
            note = text_box.get("1.0", tk.END).strip()
            test_failed_note.set(note)
            note_win.destroy()
            win.destroy()  # chiude anche la finestra principale → ritorna i dati

        tk.Button(note_win, text="Confirm", command=confirm_note).pack(pady=10)
        note_win.grab_set()  # rende la finestra modale
        note_win.wait_window()


    # --- Frame per controlli ---
    control_frame = tk.Frame(win)
    control_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)

    tk.Label(control_frame, text="PumpOff time [s]").grid(row=0, column=0, sticky="e")
    tk.Label(control_frame, text="Avg time [s]").grid(row=1, column=0, sticky="e")
    tk.Label(control_frame, text="Δt Speed [s]").grid(row=2, column=0, sticky="e")

    var_pumpOff = tk.StringVar(value=str(pumpOff_time))
    var_avgTime = tk.StringVar(value=str(avgTime))
    var_deltatime_speed = tk.StringVar(value=str(deltatime_speed))

    tk.Entry(control_frame, textvariable=var_pumpOff, width=10).grid(row=0, column=1, padx=5, pady=5)
    tk.Entry(control_frame, textvariable=var_avgTime, width=10).grid(row=1, column=1, padx=5, pady=5)
    tk.Entry(control_frame, textvariable=var_deltatime_speed, width=10).grid(row=2, column=1, padx=5, pady=5)

    button_style = {"font": ("TkDefaultFont", 10, "bold"), "width": 15, "pady": 5}
    tk.Button(control_frame, text="POST-PROCESS",command=close_window,**button_style).grid(row=3,column=0,padx=5,pady=5)
    tk.Button(control_frame, text="TEST FAILED",command=test_failed,**button_style).grid(row=4,column=0,padx=5,pady=5)
    tk.Button(control_frame, text="STOP",command=block_process,**button_style).grid(row=5,column=0,padx=5,pady=5)

    plot_tolerance_warning = tk.BooleanVar(value=plot_tolerance_warning)
# Checkbutton sotto i pulsanti
    tk.Checkbutton(control_frame,text="Plot tolerance warning",variable=plot_tolerance_warning,onvalue=True,offvalue=False).grid(row=6, column=0, columnspan=2, pady=10, sticky="w")

    # Collega l’aggiornamento automatico
    var_pumpOff.trace_add("write", update_lines)
    var_avgTime.trace_add("write", update_lines)
    var_deltatime_speed.trace_add("write", update_lines)

    # Disegna la prima volta
    update_lines()

    win.wait_window()

    return (float(var_pumpOff.get()),float(var_avgTime.get()),float(var_deltatime_speed.get()),stop_flag.get(),test_failed_note.get(),plot_tolerance_warning.get()) 


def custom_date_parser_logger(date_string):
    return pd.to_datetime(date_string, format="%d/%m/%Y %H:%M:%S.%f")


def search_change_voltage(df, val, time_delta):
    # Filtra e crea una copia indipendente del sotto-DataFrame
    time_DVpump_switch = df[(df["DVpump_Dt"] > val) | (df["DVpump_Dt"] < -val)].copy()   
    # Calcola la differenza di tempo tra righe consecutive
    time_DVpump_switch["diff_seconds"] = time_DVpump_switch["Time"].diff()
    # Rimuove eventuali NaN
    time_DVpump_switch = time_DVpump_switch.dropna(subset=["diff_seconds"])
    # Tiene solo righe dove il salto temporale è maggiore o uguale al limite imposto
    time_DVpump_switch = time_DVpump_switch[time_DVpump_switch["diff_seconds"] >= pd.Timedelta(time_delta)]
    # Converte la colonna "Time" in lista
    time_list = time_DVpump_switch["Time"].tolist()
    return time_list


def check_variazione(df, mean_value, stringa, plot_chart, tol):
    plot_chart_ = plot_chart
    range_col = df.max() - df.min()
    if df.min == 0:
        #print(stringa + " minima = 0, test fallito")
        plot_chart_ = True
        warning = "Critical"
    else:
        if range_col > 0.01 * tol * mean_value:
            #print(f"range_col {range_col} tol {tol} mean_value {mean_value}")
            #print(f"{stringa} range (max:{df.max()}, min:{df.min()}) maggiore del {tol} % della media ({mean_value}), valutare i risultati con attenzione")
            plot_chart_ = True
            warning = "Warning max:"+str(df.max())+" min:"+str(df.min())
        else:
            # print(stringa + " range minore del " + str(tol) + "% della media")
            warning = "Ok"
    return plot_chart_, warning


def files_selection():
    files = filedialog.askopenfilenames(title="Select files",filetypes=[("DEWESOFT files","*.dxd"),("CSV files", "*.csv"),("All files", "*.*")])
    if files:
        # Memorizza l’elenco completo come stringa (utile se serve in altre funzioni)
        filesread.set("; ".join(files))
        # Aggiorna anche il contatore visibile
        filescount.set(f"Files selected: {len(files)}")

        # Ricava la cartella dei file selezionati
        folder_path = os.path.dirname(files[0])
        
        # Se vuoi salvarla in una variabile Tkinter (ad es. per usarla altrove)
        selected_folder.set(folder_path)
    else:
        filescount.set("No files selected")
        selected_folder.set("")


def excel_selection():
    file = filedialog.askopenfilename(title="Select files",
                                        filetypes=[("Excel files", "*.xlsx"),("All files", "*.*")])
    if file:
        excel_file_name.set(file)

        file_parts = file.split("/")

        # Rebuild path removing leading folders until it fits
        while len(os.sep.join(file_parts)) > 50 and len(file_parts) > 1:
            file_parts.pop(0)  # remove first folder

        shorter_file_name = os.sep.join(file_parts)

        excel_file_name_short.set(shorter_file_name)
    else:
        filescount.set("No files selected")
        selected_folder.set("")


#da eliminare
def update_speed_udm(selected_option):
    # Assegna il valore alla variabile speed_udm in base alla selezione
    global speed_udm
    speed_udm = selected_option
    # print(f"Selected speed unit: {speed_udm}")

def update_speed_udm(selected_option):
    var_speed_udm.set(selected_option)  # aggiorna la variabile Tkinter


def update_filefolder(selected_option):
    # Assegna il valore alla variabile speed_udm in base alla selezione
    global filefolder_
    filefolder_ = selected_option
    button_wd.config(text=filefolder_)
    #print(f"Selected speed unit: {filefolder_}")


def toggle_labels():
    # Controlla se il flag è selezionato
    if flag_var.get():
        # Aggiungi i nuovi label se il flag è attivo
        label_tolerance_value.grid(row=8, column=0, padx=10, pady=10, sticky='w')
        label_pumpOff_time.grid(row=9, column=0, padx=10, pady=10, sticky='w')
        label_tolWarning.grid(row=10, column=0, padx=10, pady=10, sticky='w')
        entry_tolerance.grid(row=8, column=1, padx=10, pady=10)
        entry_pumpOff.grid(row=9, column=1, padx=10, pady=10)
        entry_tolWarning.grid(row=10, column=1, padx=10, pady=10)
    else:
        # Rimuovi i nuovi label se il flag è disattivato
        label_tolerance_value.grid_forget()
        label_pumpOff_time.grid_forget()
        label_tolWarning.grid_forget()
        entry_tolerance.grid_forget()
        entry_pumpOff.grid_forget()
        entry_tolWarning.grid_forget()


def import_data():
    import_data_filepath = filedialog.askopenfilename(title="Import data")
    f = open(import_data_filepath, "r")
    testo = f.readlines()
    f.close()
    for line in testo:
        if "pump name" in line.lower():
            var_pump_name.set(line.split(":")[1].replace("\n", ""))
        elif "tested voltages" in line.lower():
            var_voltage.set(line.split(":")[1].replace("\n", ""))
            var_voltage_count.set("Voltages entered:" + str(len(var_voltage.get().split(";"))))
        elif "speed unit" in line.lower():
            var_speed_udm.set( line.split(":")[1].replace("\n", ""))
        elif "tested speeds" in line.lower():
            var_speed.set(line.split(":")[1].replace("\n", ""))
            var_speed_count.set("Speeds entered:" + str(len(var_speed.get().split(";"))))
        elif "dt speeds" in line.lower():
            var_deltaTime_speed.set(line.split(":")[1].replace("\n", ""))
        elif "avg time" in line.lower():
            var_avgTime.set(line.split(":")[1].replace("\n", ""))
        elif "tolerance value" in line.lower():
            var_tolerance.set(line.split(":")[1].replace("\n", ""))
        elif "pump-off time" in line.lower():
            var_pumpOff.set(line.split(":")[1].replace("\n", ""))
        elif "tolerance warning" in line.lower():
            var_tolWarning.set(line.split(":")[1].replace("\n", ""))


def export_importdatafile():
    testo = "Pump Name: " + str(var_pump_name.get()) + "\nTested Voltages: " + str(
        var_voltage.get()) +"\nSpeed unit:"+str(var_speed_udm.get())+ "\nTested Speeds: " + str(var_speed.get()) + "\ndt speeds: " + str(
        var_deltaTime_speed.get()) + "\nAvg time: " + str(var_avgTime.get()) + "\nTolerance value: " + str(
        var_tolerance.get()) + "\nPump-off time: " + str(var_pumpOff.get()) + "\nTolerance warning: " + str(
        var_tolWarning.get())
    root = tk.Tk()
    wrk_path = filedialog.askdirectory(title="Export import data file")
    root.destroy()
    if wrk_path:
        # Ensure the file path is correct, use os.path.join for cross-platform compatibility
        file_path = os.path.join(wrk_path, "importdataexample.dat")
        # Writing to the file using a context manager (with statement)
        with open(file_path, "w") as f:
            f.writelines(testo)


def insertVOLTAGES():
    wp_window = tk.Toplevel()
    wp_window.title("Insert voltages")
    wp_window.configure(bg="lightblue")

    voltages = []

    # Treeview per mostrare i voltaggi
    tree = ttk.Treeview(wp_window, columns=("Voltage"), show="headings", height=6)
    tree.heading("Voltage", text="Voltage [V]")
    tree.column("Voltage", width=100, anchor="center")
    tree.pack(padx=10, pady=10)

    # Frame per inserimento
    entry_frame = tk.Frame(wp_window, bg="lightblue")
    entry_frame.pack(padx=10, pady=5)

    tk.Label(entry_frame, text="V:", bg="lightblue").grid(row=0, column=0, padx=5)
    voltage_entry = tk.Entry(entry_frame, width=10)
    voltage_entry.grid(row=0, column=1, padx=5)

    # Funzione per aggiungere un valore
    def add_voltage():
        try:
            vel = float(voltage_entry.get())
            voltages.append(vel)
            tree.insert("", "end", values=(vel,))
            voltage_entry.delete(0, tk.END)
        except ValueError:
            messagebox.showerror("Error", "Insert a valid number for voltage.")

    # Funzione per salvare i valori nel main
    def save_voltages():
        var_voltage.set(";".join(str(v) for v in voltages))
        var_voltage_count.set(f"Voltages entered: {len(voltages)}")
        wp_window.destroy()

    # Pulsanti
    tk.Button(entry_frame, text="Add voltage", command=add_voltage).grid(row=0, column=2, padx=10)
    tk.Button(wp_window, text="Save all voltages", command=save_voltages).pack(pady=10)

    wp_window.protocol("WM_DELETE_WINDOW", save_voltages)

def insertSPEEDS():
    wp_window = tk.Toplevel()
    wp_window.title("Insert speeds")
    wp_window.configure(bg="lightblue")

    speeds = []

    # Treeview per mostrare le velocità
    tree = ttk.Treeview(wp_window, columns=("Speed",), show="headings", height=6)
    tree.column("Speed", width=150, anchor="center")
    tree.pack(padx=10, pady=10)

    # Imposta subito il titolo colonna
    tree.heading("Speed", text="Speed " + var_speed_udm.get())

    # Frame per inserimento
    entry_frame = tk.Frame(wp_window, bg="lightblue")
    entry_frame.pack(padx=10, pady=5)

    label_speed_udm = tk.Label(entry_frame, text=var_speed_udm.get(), bg="lightblue")
    label_speed_udm.grid(row=0, column=0, padx=5)

    speed_entry = tk.Entry(entry_frame, width=10)
    speed_entry.grid(row=0, column=1, padx=5)

    # 🔄 Aggiornamento dinamico se cambia l'unità
    def update_heading(*_):
        # Se la finestra è ancora aperta e il tree esiste
        if tree.winfo_exists():
            tree.heading("Speed", text="Speed " + var_speed_udm.get())
            label_speed_udm.config(text=var_speed_udm.get())

    # Collega il trace dinamico
    trace_id = var_speed_udm.trace_add("write", update_heading)

    # Funzione per aggiungere un valore
    def add_speed():
        try:
            vel = float(speed_entry.get())
            speeds.append(vel)
            tree.insert("", "end", values=(vel,))
            speed_entry.delete(0, tk.END)
        except ValueError:
            messagebox.showerror("Error", "Insert a valid number for speed.")

    # Funzione per salvare e chiudere
    def save_speed():
        # Rimuovo il trace quando chiudo
        var_speed_udm.trace_remove("write", trace_id)
        var_speed.set(";".join(str(v) for v in speeds))
        var_speed_count.set(f"Speeds entered: {len(speeds)}")
        wp_window.destroy()

    # Pulsanti
    tk.Button(entry_frame, text="Add speed", command=add_speed).grid(row=0, column=2, padx=10)
    tk.Button(wp_window, text="Save all speeds", command=save_speed).pack(pady=10)

    wp_window.protocol("WM_DELETE_WINDOW", save_speed)


root = tk.Tk()
root.title("POST-PARAMETER SALERI")
def on_main_close():
    root.destroy()
    root.quit()  # forza la terminazione completa dell’event loop
root.protocol("WM_DELETE_WINDOW", on_main_close)
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_width = int(screen_width * 0.8)
window_height = int(screen_height * 0.8)
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2
root.geometry(f"{window_width}x{window_height}+{x}+{y}")
root.configure(bg='lightblue')

# Creiamo un frame per la sezione "Pump name" + "Select files"
frame_top = tk.Frame(root, bg="lightblue", bd=2, relief="groove")
frame_top.grid(row=0, column=0, padx=10, pady=10, sticky="w")

label_pump_name = tk.Label(frame_top, text="Insert pump name", bg="lightblue")
label_pump_name.grid(row=0, column=0, padx=10, pady=10, sticky='e')  # Posizioniamo la label in (0, 0)
# Creiamo un'entry (campo di input) in cui l'utente può scrivere una stringa
var_pump_name = tk.StringVar()
entry_pump_name = tk.Entry(frame_top, textvariable=var_pump_name)
entry_pump_name.grid(row=0, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
# Variabili dinamiche
filesread = tk.StringVar()   # contiene i nomi completi dei file
filescount = tk.StringVar()  # contiene il numero di file selezionati
excel_file_name = tk.StringVar()  # contiene il nome dell'excel selezionato
excel_file_name_short = tk.StringVar()  # contiene il nome dell'excel selezionato
selected_folder = tk.StringVar()

# Bottone per selezionare i file
button_files2read = tk.Button(frame_top, text="SELECT FILEs", command=files_selection)
button_files2read.grid(row=1, column=0, padx=10, pady=10, sticky='e')
# Label che mostra il numero dei file
label_files2read = tk.Label(frame_top, textvariable=filescount, wraplength=400, bg="lightblue")
label_files2read.grid(row=1, column=1, padx=10, pady=10, sticky='w')

# Bottone per selezionare i file
button_excel2read = tk.Button(frame_top, text="SELECT EXCEL", command=excel_selection)
button_excel2read.grid(row=1, column=2, padx=10, pady=10, sticky='e')
# Label che mostra il nome del file excel
label_files2read = tk.Label(frame_top, textvariable=excel_file_name_short, wraplength=400, bg="lightblue")
label_files2read.grid(row=1, column=3, padx=10, pady=10, sticky='w')

# Creiamo un frame BC
frame_BC = tk.LabelFrame(root, text="Boundary conditions", bg="lightblue", bd=2, relief="groove", font=("Arial", 10, "bold"))
frame_BC.grid(row=1, column=0, padx=10, pady=10, sticky="w")
# Bottone per inserire voltaggi testati
var_voltage = tk.StringVar()
var_voltage_count = tk.StringVar(value="Voltages entered: 0")
button_voltages = tk.Button(frame_BC, text="INSERT VOLTAGEs", command=insertVOLTAGES)
button_voltages.grid(row=0, column=0, padx=10, pady=10, sticky='e')
label_voltages=tk.Label(frame_BC, textvariable=var_voltage_count , wraplength=400, bg="lightblue")
label_voltages.grid(row=0, column=1, padx=10, pady=10, sticky='e')
# Label udm Speed
label_udm_speed=tk.Label(frame_BC,text="Speed unit:",bg="lightblue")
label_udm_speed.grid(row=1, column=0, padx=10, pady=10, sticky='e')
options = ["rpm", "pwm", "lin"]
var_speed_udm = tk.StringVar()
var_speed_udm.set("")
dropdown_speed = tk.OptionMenu(frame_BC, var_speed_udm, *options, command=update_speed_udm)
dropdown_speed.grid(row=1, column=1, padx=10, pady=10)
#Bottone per inserire le velocità testate
var_speed=tk.StringVar()
var_speed_count=tk.StringVar(value="Speeds entered: 0")
button_speeds = tk.Button(frame_BC, text="INSERT SPEEDs", command=insertSPEEDS)
button_speeds.grid(row=2, column=0, padx=10, pady=10, sticky='e')
label_speeds=tk.Label(frame_BC, textvariable=var_speed_count , wraplength=400, bg="lightblue")
label_speeds.grid(row=2, column=1, padx=10, pady=10, sticky='e')
# Creiamo label per definire i voltaggi testati
label_deltaTime_speed = tk.Label(frame_BC, text="dt Speed [s]", bg="lightblue")
label_deltaTime_speed.grid(row=3, column=0, padx=10, pady=10, sticky='e')
# Creiamo variabile dinamica per la stringa del tempo per ogni velocità
var_deltaTime_speed = tk.StringVar()
# Creiamo un'entry (campo di input) in cui l'utente può scrivere una stringa
entry_deltaTime_speed = tk.Entry(frame_BC, textvariable=var_deltaTime_speed)
entry_deltaTime_speed.grid(row=3, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
label_deltaTime_speed_info = tk.Label(frame_BC, text="*Stationary time for each speed", bg="lightblue")
label_deltaTime_speed_info.grid(row=3, column=2, padx=10, pady=10, sticky='e')
# Creiamo label per definire i voltaggi testati
label_avgTime = tk.Label(frame_BC, text="Avg time [s]", bg="lightblue")
label_avgTime.grid(row=4, column=0, padx=10, pady=10, sticky='e')
# Creiamo variabile dinamica per la stringa della working directory
var_avgTime = tk.StringVar()
# Creiamo un'entry (campo di input) in cui l'utente può scrivere una stringa
entry_avgTime = tk.Entry(frame_BC, textvariable=var_avgTime)
entry_avgTime.grid(row=4, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
label_avgTime_info = tk.Label(frame_BC, text="*Average time for each speed", bg="lightblue")
label_avgTime_info.grid(row=4, column=2, padx=10, pady=10, sticky='e')

# Creiamo un frame More Options
frame_MoreOptions = tk.LabelFrame(root, text="More Options", bg="lightblue", bd=2, relief="groove", font=("Arial", 10, "bold"))
frame_MoreOptions.grid(row=2, column=0, padx=10, pady=10, sticky="w")
# Nuovi label da mostrare quando il checkbox è selezionato
label_tolerance_value = tk.Label(frame_MoreOptions, text="Tolerance value dV/dt", bg="lightblue")
label_tolerance_value.grid(row=0, column=0, padx=10, pady=10, sticky='e')
label_pumpOff_time = tk.Label(frame_MoreOptions, text="Pump off time [s]", bg="lightblue")
label_pumpOff_time.grid(row=1, column=0, padx=10, pady=10, sticky='e')
label_tolWarning = tk.Label(frame_MoreOptions, text="Tolerance warning %", bg="lightblue")
label_tolWarning.grid(row=2, column=0, padx=10, pady=10, sticky='e')
var_tolerance = tk.StringVar()
var_tolerance.set("5")
var_pumpOff = tk.StringVar()
var_pumpOff.set("30")
var_tolWarning = tk.StringVar()
var_tolWarning.set("2.5")
entry_tolerance = tk.Entry(frame_MoreOptions, textvariable=var_tolerance)
entry_tolerance.grid(row=0, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
entry_pumpOff = tk.Entry(frame_MoreOptions, textvariable=var_pumpOff)
entry_pumpOff.grid(row=1, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label
entry_tolWarning = tk.Entry(frame_MoreOptions, textvariable=var_tolWarning)
entry_tolWarning.grid(row=2, column=1, padx=10, pady=10)  # Posizioniamo l'entry a destra della label


# Creiamo un frame Buttons
frame_Buttons = tk.LabelFrame(root, text="Functions", bg="lightblue", bd=2, relief="groove", font=("Arial", 10, "bold"))
frame_Buttons.grid(row=2, column=2, padx=10, pady=10, sticky="w")

# Creiamo un comando di import data file
button1 = tk.Button(frame_Buttons, text="Import data", relief="ridge", command=lambda: [import_data()])
button1.grid(row=0, column=0, padx=10, pady=10)
# Creiamo un comando di export import data
button2 = tk.Button(frame_Buttons, text="Export Import data file", relief="ridge", command=lambda: [export_importdatafile()])
button2.grid(row=0, column=3, padx=10, pady=10)
# Creiamo un pulsante per lanciare il post-processing
button = tk.Button(frame_Buttons, text="EXECUTE", command=lambda: [post_parameter()], width=20, height=3, font=("Arial", 14))
button.grid(row=1, column=1, padx=10, pady=10)


# Avviamo il loop principale della finestra
root.mainloop()


