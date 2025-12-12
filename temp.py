from datetime import datetime

from PySide6.QtCore import QTimer
from PySide6.QtWidgets import (
    QApplication, QWidget, QPushButton, QLabel,
    QHBoxLayout, QVBoxLayout, QFileDialog, QMessageBox
)
from PySide6.QtGui import QIcon
import pandas as pd
import os


class Automatic_Mode_Button(QPushButton):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCheckable(True)
        self.setChecked(False)
        self.update_style()
        self.toggled.connect(self.update_style)

    def update_style(self):
        if self.isChecked():
            self.setText("AUTOMATIC MODE")
            self.setStyleSheet("""
                QPushButton {
                    background-color: #4cd964;
                    color: white;
                    border-radius: 14px;
                    padding: 6px 12px;
                }
            """)
        else:
            self.setText("MANUAL MODE")
            self.setStyleSheet("""
                QPushButton {
                    background-color: #ff3131;
                    color: white;
                    border-radius: 14px;
                    padding: 6px 12px;
                }
            """)

    def automatic_mode_active(self):
        return True if self.isChecked() else False

class Window(QWidget):
    def __init__(self):
        super().__init__()

        # Switch
        self.switch = Automatic_Mode_Button()
        self.switch.clicked.connect(self.automatic_button_action)

        # Bottone Load Excel
        self.btn_load_excel = QPushButton("Load Excel")
        self.btn_load_excel.setIcon(QIcon.fromTheme("document-open"))
        self.btn_load_excel.clicked.connect(self.load_excel)
        self.btn_load_excel.setEnabled(False)

        # Stile simile a Start/Stop
        self.btn_load_excel.setStyleSheet("""
            QPushButton {
                background-color: #007bff;  /* blu */
                color: white;
                border-radius: 10px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover:!disabled { background-color: #0069d9; }
            QPushButton:disabled { background-color: #a0a0a0; color: #666666; }
        """)


        # Label per mostrare il nome del file
        self.label_file = QLabel("Nessun file selezionato")

        # Start e Stop
        self.btn_start = QPushButton("▶ Start")
        self.btn_stop = QPushButton("■ Stop")
        self.btn_start.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border-radius: 10px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #218838; }
            QPushButton:disabled { background-color: #a0a0a0; color: #666666; }
        """)
        self.btn_stop.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border-radius: 10px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #c82333; }
            QPushButton:disabled { background-color: #a0a0a0; color: #666666; }

        """)
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(False)
        # Collega Start e Stop
        self.btn_start.clicked.connect(self.start_action)
        self.btn_stop.clicked.connect(self.stop_action)

        # --- Layout superiore ---
        top_row = QHBoxLayout()
        top_row.addWidget(self.switch)

        # --- Layout inferiore ---
        bottom_row = QHBoxLayout()
        bottom_row.addWidget(self.btn_load_excel)
        bottom_row.addWidget(self.label_file)
        bottom_row.addWidget(self.btn_start)
        bottom_row.addWidget(self.btn_stop)

        # Layout principale
        main = QVBoxLayout()
        main.addLayout(top_row)
        main.addLayout(bottom_row)
        self.setLayout(main)
        self.setWindowTitle("Switch & Bottoni Stile")

        # Timer per il "player"
        self.timer = QTimer()
        self.timer.timeout.connect(self.next_row)
        self.current_index = 0
        self.data = []

    def automatic_button_action(self):
        if self.switch.automatic_mode_active():
            self.btn_load_excel.setDisabled(False)
            self.btn_start.setDisabled(True)
            self.btn_stop.setDisabled(True)
        else:
            self.btn_load_excel.setDisabled(True)
            self.btn_start.setDisabled(True)
            self.btn_stop.setDisabled(True)
    def load_excel(self):
        # Apri finestra dialogo per Excel
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            try:
                from openpyxl import load_workbook

                # Carica il file Excel
                wb = load_workbook(file_path)  # percorso del file
                sheet = wb.active  # usa il primo foglio attivo

                # Leggi il valore della cella A1
                first_row_parameter = sheet["A1"].value
                first_row_value = sheet["B1"].value
                first_row_delay = sheet["C1"].value

                # Controlla prima cella
                if (first_row_parameter == "Parameter" and
                        first_row_value == "Value" and
                        first_row_delay == "Delay[s]"):
                    # Aggiorna label con il nome del file
                    file_name = os.path.basename(file_path)
                    self.label_file.setText(file_name)
                    self.data = []  # lista per contenere tutte le righe non vuote

                    for row in sheet.iter_rows(min_row=1, values_only=True):
                        # Controlla se la riga contiene almeno un valore
                        if any(cell is not None for cell in row):
                            parameter = row[0] if len(row) > 1 else ""
                            value = row[1] if len(row) > 2 else ""
                            delay_s = row[2]
                            self.data.append((parameter, value, delay_s))

                    verbose = True
                    if verbose:
                        # Stampa tutte le righe
                        print("Dati non vuoti:")
                        for r in self.data:
                            print(r)

                    # Stampa numero di righe non vuote
                    QMessageBox.information(self, "Excel Caricato", f"File caricato correttamente\n"
                                                                    f"Numero di step di test: {len(self.data) - 1}")
                    self.btn_start.setEnabled(True)

                else:
                    QMessageBox.critical(self, "Errore",
                                         "Il file Excel non è formattato correttamente")

            except Exception as e:
                QMessageBox.critical(self, "Errore", f"Impossibile leggere il file:\n{e}")

    def start_action(self):
        # Abilita Stop e disabilita Start
        self.btn_stop.setEnabled(True)
        self.btn_start.setEnabled(False)
        self.start_time = 0  # riferimento tempo iniziale
        self.timer.start(100)  # timer veloce, calcola ritardo dai tempi
        print("Start premuto")
        print(self.data)
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-5]  # hh:mm:ss.mmm
        print(f"[{timestamp}] Test started:")
        self.current_index = 1
        self.timer.start(100)

    def stop_action(self):
        self.timer.stop()
        # Disabilita Stop e abilita Start
        self.btn_stop.setEnabled(False)
        self.btn_start.setEnabled(True)
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-5]  # hh:mm:ss.mmm
        print(f"[{timestamp}] Test completed:")


    def next_row(self):
        # Stampa con timestamp corrente
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-5]  # hh:mm:ss.mmm

        if self.current_index < len(self.data):
            string1, string2, _ = self.data[self.current_index]
            print(f"[{timestamp}] Parameter: {string1}, Value: {string2}")

            # Calcolo intervallo per la prossima riga
            if self.current_index < len(self.data) - 1:
                delay = self.data[self.current_index][2]
                interval = max(0, delay * 1000)  # converti in ms
                self.timer.start(interval)

            self.current_index += 1
        else:
            self.timer.stop()
            self.btn_stop.setEnabled(False)
            self.btn_start.setEnabled(True)
            print(f"[{timestamp}] Test completed")
            return


if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    win = Window()
    win.show()
    sys.exit(app.exec())
